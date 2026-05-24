# -*- coding: utf-8 -*-
"""
Created on Mon Jun 17 14:27:40 2024

@author: cwtian
"""
import os
import numpy as np
import pandas as pd
from copy import deepcopy
from itertools import product
from openpyxl.styles import PatternFill
import time

from .models import MIPError

class CONST(object):
    NORMAL='normal'
    REVISIONAL='revisional'
    BACK='back'
    CREATED='created'
    INPROCESS='inprocess'
    SUCCESS='success'
class MaxDistanceError(Exception):
    def __init__(self,message,df=None,path=None):
        if df!=None and path!=None:
            with pd.ExcelWriter(path) as writer:
                df.to_excel(writer)
        self.message=message
    def __str__(self):
        return self.message
class heuristic(object):
    def __init__(self,situation=CONST.NORMAL,*args,**kargs):
        if situation==CONST.NORMAL:
            self.__instance=heuristic_normal(*args,**kargs)
        elif situation==CONST.REVISIONAL:
            self.__instance=heuristic_revisional(*args,**kargs)
        elif situation==CONST.BACK:
            self.__instance=heuristic_back(*args,**kargs)
    def run(self):
        self.__instance.run()
    def schedule_frame(self):
        return self.__instance.schedule_frame()
class heuristic_normal(object):
    def __init__(self,workPath=None,decimal=5,
                 infoPath='info.xlsx',taskPath='task.xlsx',packPath='package.xlsx',
                 pointPath='point.xlsx',distancePath='distance.xlsx',timePath='time.xlsx',
                 powerPath='power.xlsx',outputPath='schedule.xlsx',
                 dataFrames=None,writeOutput=True):
        print('Process initiated')
        self.__starttime=time.time()
        self.__decimal=decimal
        if workPath==None:
            self.__infoPath,self.__taskPath,self.__pakpath=infoPath,taskPath,packPath
            self.__pointpath,self.__distancepath,self.__timepath,self.__powerpath=pointPath,distancePath,timePath,powerPath
        else:
            self.__infoPath,self.__taskPath,self.__pakpath=os.path.join(workPath,infoPath),os.path.join(workPath,taskPath),os.path.join(workPath,packPath)
            self.__pointpath,self.__distancepath,self.__timepath,self.__powerpath=os.path.join(workPath,pointPath),os.path.join(workPath,distancePath),os.path.join(workPath,timePath),os.path.join(workPath,powerPath)
        self.__outputpath=outputPath
        self.__dataframes=dataFrames or {}
        self.__write_output=writeOutput
        self.__Status=CONST.CREATED
    def test_IO(self):
        if self.__dataframes:
            missing=[key for key in ['info','task','package','point','distance','time','power'] if key not in self.__dataframes]
            if missing:
                raise ValueError(f'Missing DataFrame inputs: {missing}')
            return None
        FileNotFound=[]
        if not os.path.exists(self.__infoPath):
            FileNotFound.append(self.__infoPath)
        if not os.path.exists(self.__taskPath):
            FileNotFound.append(self.__taskPath)
        if not os.path.exists(self.__pakpath):
            FileNotFound.append(self.__pakpath)
        if not os.path.exists(self.__pointpath):
            FileNotFound.append(self.__pointpath)
        if not os.path.exists(self.__distancepath):
            FileNotFound.append(self.__distancepath)
        if not os.path.exists(self.__timepath):
            FileNotFound.append(self.__timepath)
        if not os.path.exists(self.__powerpath):
            FileNotFound.append(self.__powerpath)
        if FileNotFound!=[]:
            raise FileNotFoundError('No such file: {}'.format(' '.join(FileNotFound)))
        notPermitted=False
        if self.__write_output and os.path.exists(self.__outputpath):
            try:
                pd.read_excel(self.__outputpath)
            except PermissionError:
                notPermitted=True
        if notPermitted:
            raise PermissionError(f'Permission denied: {self.__outputpath}')
        return None
    def read_info(self):
        info=self.__dataframes.get('info')
        info=info.copy() if info is not None else pd.read_excel(self.__infoPath)
        self.__MDistance=info['max-distance'][0]
        self.__TTime=list(map(float,info['total-time/day'][0].split(';')))
        self.__TPower=list(map(float,info['total-power/day'][0].split(';')))
        self.__Mincontinuous=info['min-continuous'][0]
        self.__12gap=info['12-gap'][0]
        self.__23gap=info['23-gap'][0]
        return None
    def __read_matrix(self,path,key):
        matrix=self.__dataframes.get(key)
        if matrix is not None:
            matrix=matrix.copy()
        else:
            matrix=pd.read_excel(path)
            matrix.set_index(matrix.columns[0],inplace=True)
            matrix.index.rename(None,inplace=True)
        return matrix
    def read_task(self):
        task=self.__dataframes.get('task')
        self.__task=task.copy() if task is not None else pd.read_excel(self.__taskPath)
        # print(self.__task.head())
        self.__task['No']=range(self.__task.shape[0])
        self.__dmatrix=self.__read_matrix(self.__distancepath,'distance')
        self.__dmatrix.replace(np.inf,self.__MDistance*3,inplace=True)
        np.fill_diagonal(self.__dmatrix.values,0)
        self.__tmatrix=self.__read_matrix(self.__timepath,'time')
        self.__tmatrix.replace(np.inf,max(self.__TTime),inplace=True)
        np.fill_diagonal(self.__tmatrix.values,0)
        self.__pmatrix=self.__read_matrix(self.__powerpath,'power')
        self.__pmatrix.replace(np.inf,max(self.__TPower),inplace=True)
        np.fill_diagonal(self.__pmatrix.values,0)
        return None
    def read_package(self):
        package=self.__dataframes.get('package')
        self.__package=package.copy() if package is not None else pd.read_excel(self.__pakpath)
        time=[sum(self.__package[self.__package['tag']==tag]['time']) for tag in ['D1ss','D1se','D1es','D1ee','D2ss','D2se','D2es','D2ee','D3ss','D3se','D3es','D3ee']]
        power=[sum(self.__package[self.__package['tag']==tag]['power']) for tag in ['D1ss','D1se','D1es','D1ee','D2ss','D2se','D2es','D2ee','D3ss','D3se','D3es','D3ee']]
        self.__PKGTime=[time[0]+time[1]+time[2]+time[3],time[4]+time[5]+time[6]+time[7],time[8]+time[9]+time[10]+time[11]]
        self.__PKGPower=[power[0]+power[1]+power[2]+power[3],power[4]+power[5]+power[6]+power[7],power[8]+power[9]+power[10]+power[11]]
        self.__TTime=[self.__TTime[i]-self.__PKGTime[i] for i in range(3)]
        self.__TPower=[self.__TPower[i]-self.__PKGPower[i] for i in range(3)]
        return None
    def read_point(self):
        point=self.__dataframes.get('point')
        self.__pointdf=point.copy() if point is not None else pd.read_excel(self.__pointpath)
        self.__pointdf.set_index(self.__pointdf.columns[0],inplace=True)
        self.__pointdf.index.rename(None,inplace=True)
        return None
    def __add_void_matrix(self,matrix):
        voidrow=[matrix.loc['探测起点',:].values.tolist() for i in range(4)]
        voidrowdf=pd.DataFrame(voidrow,index=['探测起点1','探测起点2','探测起点3','探测起点4'],columns=matrix.columns)
        matrix=pd.concat([voidrowdf,matrix])
        matrix.drop(['探测起点'],inplace=True)
        voidcol=np.array([matrix.loc[:,'探测起点'].values.tolist() for i in range(4)]).T
        voidcoldf=pd.DataFrame(voidcol,columns=['探测起点1','探测起点2','探测起点3','探测起点4'],index=matrix.index)
        matrix=pd.concat([voidcoldf,matrix],axis=1)
        matrix.drop(['探测起点'],axis=1,inplace=True)
        for i,j in product(['探测起点1','探测起点2','探测起点3','探测起点4'],['探测起点1','探测起点2','探测起点3','探测起点4']):
            if i!=j:
                matrix.loc[i,j]=0
        return matrix
    def gen_void_point(self):
        voidpoint=pd.DataFrame({'X':[self.__pointdf.loc['探测起点','X'],self.__pointdf.loc['探测起点','X'],self.__pointdf.loc['探测起点','X'],self.__pointdf.loc['探测起点','X']],
                                'Y':[self.__pointdf.loc['探测起点','Y'],self.__pointdf.loc['探测起点','Y'],self.__pointdf.loc['探测起点','Y'],self.__pointdf.loc['探测起点','Y']],
                                '备注':['虚拟原点：第一天出发','虚拟原点：第一天返回第二天出发','虚拟原点：第二天返回第三天出发','虚拟原点：第三天返回']},
                               index=['探测起点1','探测起点2','探测起点3','探测起点4'])
        self.__pointdf=pd.concat([voidpoint,self.__pointdf])
        self.__pointdf.drop(['探测起点'],inplace=True)
        self.__pointdf['No']=range(self.__pointdf.shape[0])
        self.__dmatrix=self.__add_void_matrix(self.__dmatrix)
        self.__tmatrix=self.__add_void_matrix(self.__tmatrix)
        self.__pmatrix=self.__add_void_matrix(self.__pmatrix)
        self.__task['location']=self.__task['location'].replace('探测起点','探测起点1,探测起点2,探测起点3,探测起点4')
        self.__task['location']=self.__task['location'].fillna(','.join(self.__pointdf.index.values))
        location=self.__task['location'].values
        pool=[]
        for i in location:
            if pd.isna(i):
                pool.append(np.nan)
            else:
                split=list(i.split(','))
                pool.append(split)
        self.__task['location']=pd.Series(pool)
        voidtaskdf=pd.DataFrame({'No':[self.__task.shape[0],self.__task.shape[0]+1,self.__task.shape[0]+2,self.__task.shape[0]+3],
                                 'name':['void1','void2','void3','void4'],
                                 'revenue':[0,0,0,0],
                                 'location':[['探测起点1'],['探测起点2'],['探测起点3'],['探测起点4']],
                                 'day':[np.nan,np.nan,np.nan,np.nan],
                                 'time':[0,0,0,0],
                                 'power':[0,0,0,0],
                                 'required':[False,False,False,False],
                                 'continuous':[False,False,False,False],
                                 'remote':[False,False,False,False],
                                 'exceptO':[False,False,False,False],
                                 'tag':[np.nan,np.nan,np.nan,np.nan]},
                                index=[self.__task.shape[0],self.__task.shape[0]+1,self.__task.shape[0]+2,self.__task.shape[0]+3])
        self.__task=pd.concat([self.__task,voidtaskdf])
        self.__opoint=((self.__pointdf.loc['探测起点1','No'],self.__task[self.__task['name']=='void1'].index[0]),
                       (self.__pointdf.loc['探测起点2','No'],self.__task[self.__task['name']=='void2'].index[0]),
                       (self.__pointdf.loc['探测起点3','No'],self.__task[self.__task['name']=='void3'].index[0]),
                       (self.__pointdf.loc['探测起点4','No'],self.__task[self.__task['name']=='void4'].index[0]))
        self.__pointdfbackup=self.__pointdf.copy()
        self.__pointdfbackup.reset_index(inplace=True)
        self.__pointdfbackup.set_index('No',inplace=True)
        return None
    def __cartesian_to_polar(self,x,y,O):
        r=np.sqrt((x-O[0])**2+(y-O[1])**2)
        theta=np.arctan2(y-O[1],x-O[0])
        return r,theta
    def __polar_to_cartesian(self,r,theta,O):
        x=r*np.cos(theta)+O[0]
        y=r*np.sin(theta)+O[1]
        return x,y
    def __cal_distance(self,point1,point2):
        dis=((point1[0]-point2[0])**2+(point1[1]-point2[1])**2)**0.5
        return dis
    def __from_distance(self,distance):
        pt=(distance,0)
        return pt
    def check_remote(self):
        remoteindex=self.__task[self.__task['remote']==True].index.to_list()
        points=[]
        for index in remoteindex:
            pts=[]
            for point in self.__task.loc[index,'location']:
                if self.__dmatrix.loc['探测起点1',point]>=self.__MDistance:
                    points.append(point)
                    pts.append(point)
            self.__task.at[index,'location']=pts
        if len(points)==0:
            start=0
            path='new-point.xlsx'
            sortedseries=self.__dmatrix.sort_values(by=['探测起点1'],ascending=False).loc['探测起点1',:]
            while sortedseries[start]>=self.__MDistance:
                start+=1
            pts=self.__dmatrix.sort_values(by=['探测起点1'],ascending=False).iloc[start:start+5,:].index.values
            r,_=self.__cartesian_to_polar(self.__from_distance(self.__MDistance)[0],self.__from_distance(self.__MDistance)[1],(0,0))
            pts=[self.__cartesian_to_polar(self.__pointdf.loc[pt,'X'],self.__pointdf.loc[pt,'Y'],(self.__pointdf.loc['探测起点1','X'],self.__pointdf.loc['探测起点1','Y'])) for pt in pts]
            pts=[self.__polar_to_cartesian(r,pt[1],(self.__pointdf.loc['探测起点1','X'],self.__pointdf.loc['探测起点1','Y'])) for pt in pts]
            newpt=pd.DataFrame(pts,columns=['X','Y'],index=pd.Index(['新最远探测点1','新最远探测点2','新最远探测点3','新最远探测点4','新最远探测点5'],name='name'))
            raise MaxDistanceError(f'No point meets the max distance requirement. Recommended points have been generated in {path}',newpt,path)
        return None
    def divide_task(self):
        self.__reqtaskindex=self.__task[self.__task['required']==True].index.to_list()
        self.__opttaskindex=self.__task[self.__task['required']==False].index.to_list()
        self.__daytaskindex=[self.__task[self.__task['day']==1].index.to_list(),
                             self.__task[self.__task['day']==2].index.to_list(),
                             self.__task[self.__task['day']==3].index.to_list(),
                             self.__task[self.__task['day']=='1,2'].index.to_list(),
                             self.__task[self.__task['day']=='2,3'].index.to_list(),
                             self.__task[self.__task['day']=='1,3'].index.to_list()]
        self.__remtaskindex=self.__task[self.__task['remote']==True].index.to_list()
        self.__tagtaskindex=[self.__task[self.__task['tag']=='12s'].index.to_list(),
                             self.__task[self.__task['tag']=='12e'].index.to_list(),
                             self.__task[self.__task['tag']=='23s'].index.to_list(),
                             self.__task[self.__task['tag']=='23e'].index.to_list()]
        self.__contaskindex=self.__task[self.__task['continuous']==True].index.to_list()
        tagtask=[]
        for i in range(len(self.__tagtaskindex)):
            tagtask.append(*self.__tagtaskindex[i])
        for index in self.__reqtaskindex.copy():
            if index in self.__contaskindex:
                self.__reqtaskindex.remove(index)
            if index in tagtask:
                self.__reqtaskindex.remove(index)
        for index in self.__opttaskindex.copy():
            if index in self.__contaskindex:
                self.__opttaskindex.remove(index)
            if index in tagtask:
                self.__opttaskindex.remove(index)
            if index in [self.__opoint[0][1],self.__opoint[1][1],self.__opoint[2][1],self.__opoint[3][1]]:
                self.__opttaskindex.remove(index)
        self.__noOtaskindex=self.__task[self.__task['exceptO']==True].index.to_list()
        return None
    def drop_O(self):
        for i in self.__noOtaskindex:
            self.__task.loc[i,'location'].remove('探测起点1')
            self.__task.loc[i,'location'].remove('探测起点2')
            self.__task.loc[i,'location'].remove('探测起点3')
            self.__task.loc[i,'location'].remove('探测起点4')
        return None
    def __check_plan(self,plan,day,node):
        time,power=self.__eval_plan_time_power(plan)
        timeOK=time[0]<=self.__TTimeRev[0] and time[1]<=self.__TTimeRev[1] and time[2]<=self.__TTimeRev[2]
        powerOK=power[0]<=self.__TPowerRev[0] and power[1]<=self.__TPowerRev[1] and power[2]<=self.__TPowerRev[2]
        backtime,backpower=self.__eval_back_time_power(plan,day,node)
        backtimeOK=backtime[0]<=self.__TTimeRev[0] and backtime[1]<=self.__TTimeRev[1] and backtime[2]<=self.__TTimeRev[2]
        backpowerOK=backpower[0]<=self.__TPowerRev[0] and backpower[1]<=self.__TPowerRev[1] and backpower[2]<=self.__TPowerRev[2]
        return timeOK and powerOK and backtimeOK and backpowerOK
    def __eval_plan_time_power(self,plan):
        time_usage,power_usage=[0.,0.,0.],[0.,0.,0.]
        i=0
        curnode=None
        for dayplan in plan:
            for node in dayplan:
                if curnode==None:
                    curnode=node
                else:
                    if curnode[0]==node[0]:
                        time=self.__task.loc[node[1],'time']
                        power=self.__task.loc[node[1],'power']
                    else:
                        time=self.__tmatrix.loc[self.__pointdfbackup.loc[curnode[0],'index'],self.__pointdfbackup.loc[node[0],'index']]+self.__task.loc[node[1],'time']
                        power=self.__pmatrix.loc[self.__pointdfbackup.loc[curnode[0],'index'],self.__pointdfbackup.loc[node[0],'index']]+self.__task.loc[node[1],'power']
                    time_usage[i]+=time
                    power_usage[i]+=power
                    curnode=node
            i+=1
        return time_usage,power_usage
    def __eval_back_time_power(self,plan,day,newnode):
        time_usage,power_usage=[0.,0.,0.],[0.,0.,0.]
        curnode=None
        end=False
        for node in plan[day]:
            if curnode==None:
                curnode=node
            elif curnode==newnode:
                time=self.__tmatrix.loc[self.__pointdfbackup.loc[curnode[0],'index'],self.__pointdfbackup.loc[self.__opoint[day+1][0],'index']]
                power=self.__pmatrix.loc[self.__pointdfbackup.loc[curnode[0],'index'],self.__pointdfbackup.loc[self.__opoint[day+1][0],'index']]
                end=True
            else:
                if curnode[0]==node[0]:
                    time=self.__task.loc[node[1],'time']
                    power=self.__task.loc[node[1],'power']
                else:
                    time=self.__tmatrix.loc[self.__pointdfbackup.loc[curnode[0],'index'],self.__pointdfbackup.loc[node[0],'index']]+self.__task.loc[node[1],'time']
                    power=self.__pmatrix.loc[self.__pointdfbackup.loc[curnode[0],'index'],self.__pointdfbackup.loc[node[0],'index']]+self.__task.loc[node[1],'power']
                time_usage[day]+=time
                power_usage[day]+=power
                curnode=node
                if end:
                    break
        return time_usage,power_usage
    def __cal_score(self,plan,newnode,newplan,w=(1.,0.)):
        time,power=self.__eval_plan_time_power(plan)
        newtime,newpower=self.__eval_plan_time_power(newplan)
        score=w[0]*(self.__task.loc[newnode[1],'revenue']/(sum(newtime)-sum(time)))+w[1]*(self.__task.loc[newnode[1],'revenue']/(sum(newpower)-sum(power)))
        return score
    def __find_best_pos(self,plan,node):
        bplan=None
        btaskno=None
        bscore=0.
        if node[1] in self.__daytaskindex[0]:
            days=(0,)
        elif node[1] in self.__daytaskindex[1]:
            days=(1,)
        elif node[1] in self.__daytaskindex[2]:
            days=(2,)
        elif node[1] in self.__daytaskindex[3]:
            days=(0,1)
        elif node[1] in self.__daytaskindex[4]:
            days=(1,2)
        elif node[1] in self.__daytaskindex[5]:
            days=(0,2)
        else:
            days=(0,1,2)
        for day in days:
            for i in range(1,len(plan[day])):
                newplan=deepcopy(plan)
                newplan[day].insert(i,node)
                if self.__check_plan(newplan,day,node):
                    score=self.__cal_score(plan,node,newplan)
                else:
                    score=0.
                if score>bscore:
                    bscore=score
                    bplan=deepcopy(newplan)
                    btaskno=node[1]
        return bplan,bscore,btaskno
    def __iter_task(self,plan,tasklist):
        plancopy=deepcopy(plan)
        bbplan=None
        bbtaskno=None
        bbscore=0.
        nodes=[]
        for taskno in tasklist:
            pts=self.__task.loc[taskno,'location']
            for pt in pts:
                ptno=self.__pointdf.loc[pt,'No']
                nodes.append((ptno,taskno))
        for node in nodes:
            bplan,bscore,btaskno=self.__find_best_pos(plancopy,node)
            if bscore>bbscore:
                bbscore=bscore
                bbplan=deepcopy(bplan)
                bbtaskno=btaskno
        return bbplan,bbtaskno
    def _gen_init_plan(self,n=None):
        if n==None:
            print('Solving feasible plan')
        else:
            print(f'Solving feasible plan for Occasion {n}')
        self.__Status=CONST.INPROCESS
        self.__res=[[self.__opoint[0],self.__opoint[1]],[self.__opoint[1],self.__opoint[2]],[self.__opoint[2],self.__opoint[3]]]
        return None
    def _proc_tag(self,days=(0,1)):
        if self.__Status==CONST.INPROCESS:
            self.__TTimeRev=deepcopy(self.__TTime)
            self.__TPowerRev=deepcopy(self.__TPower)
            if days[0]==0:
                self.__TTimeRev[0]=self.__TTimeRev[0]-self.__task.loc[self.__task[self.__task['tag']=='12s'].index[0],'time']-self.__task.loc[self.__task[self.__task['tag']=='12e'].index[0],'time']
                self.__TPowerRev[0]=self.__TPowerRev[0]-self.__task.loc[self.__task[self.__task['tag']=='12s'].index[0],'power']-self.__task.loc[self.__task[self.__task['tag']=='12e'].index[0],'power']
            elif days[0]==1:
                self.__TTimeRev[1]=self.__TTimeRev[1]-self.__task.loc[self.__task[self.__task['tag']=='12s'].index[0],'time']-self.__task.loc[self.__task[self.__task['tag']=='12e'].index[0],'time']
                self.__TPowerRev[1]=self.__TPowerRev[1]-self.__task.loc[self.__task[self.__task['tag']=='12s'].index[0],'power']-self.__task.loc[self.__task[self.__task['tag']=='12e'].index[0],'power']
            if days[1]==1:
                self.__TTimeRev[1]=self.__TTimeRev[1]-self.__task.loc[self.__task[self.__task['tag']=='23s'].index[0],'time']-self.__task.loc[self.__task[self.__task['tag']=='23e'].index[0],'time']
                self.__TPowerRev[1]=self.__TPowerRev[1]-self.__task.loc[self.__task[self.__task['tag']=='23s'].index[0],'power']-self.__task.loc[self.__task[self.__task['tag']=='23e'].index[0],'power']
            elif days[1]==2:
                self.__TTimeRev[2]=self.__TTimeRev[2]-self.__task.loc[self.__task[self.__task['tag']=='23s'].index[0],'time']-self.__task.loc[self.__task[self.__task['tag']=='23e'].index[0],'time']
                self.__TPowerRev[2]=self.__TPowerRev[2]-self.__task.loc[self.__task[self.__task['tag']=='23s'].index[0],'power']-self.__task.loc[self.__task[self.__task['tag']=='23e'].index[0],'power']
        return None
    def _proc_cont(self):
        if self.__Status==CONST.INPROCESS:
            contask=deepcopy(self.__contaskindex)
            i=1
            while i<=self.__Mincontinuous:
                plan,taskno=self.__iter_task(self.__res,contask)
                if plan==None:
                    raise MIPError('Cannot assign all continuous tasks')
                else:
                    self.__res=plan
                    contask.remove(taskno)
                    i+=1
        return None
    def _proc_req(self):
        if self.__Status==CONST.INPROCESS:
            reqtask=deepcopy(self.__reqtaskindex)
            while len(reqtask)>0:
                plan,taskno=self.__iter_task(self.__res,reqtask)
                if plan==None:
                    raise MIPError('Cannot assign all required tasks')
                else:
                    self.__res=plan
                    reqtask.remove(taskno)
        return None
    def _add_tag(self,days=(0,1)):
        if self.__Status==CONST.INPROCESS:
            if days[0]==0:
                self.__res[0].insert(1,(self.__pointdf.loc['探测起点1','No'],self.__task[self.__task['tag']=='12s'].index[0]))
                self.__res[0].insert(-1,(self.__pointdf.loc['探测起点2','No'],self.__task[self.__task['tag']=='12e'].index[0]))
            elif days[0]==1:
                self.__res[1].insert(1,(self.__pointdf.loc['探测起点1','No'],self.__task[self.__task['tag']=='12s'].index[0]))
                self.__res[1].insert(-1,(self.__pointdf.loc['探测起点2','No'],self.__task[self.__task['tag']=='12e'].index[0]))
            if days[1]==1:
                self.__res[1].insert(1,(self.__pointdf.loc['探测起点2','No'],self.__task[self.__task['tag']=='23s'].index[0]))
                self.__res[1].insert(-1,(self.__pointdf.loc['探测起点3','No'],self.__task[self.__task['tag']=='23e'].index[0]))
            elif days[1]==2:
                self.__res[2].insert(1,(self.__pointdf.loc['探测起点2','No'],self.__task[self.__task['tag']=='23s'].index[0]))
                self.__res[2].insert(-1,(self.__pointdf.loc['探测起点3','No'],self.__task[self.__task['tag']=='23e'].index[0]))
        return None
    def _proc_opt(self):
        if self.__Status==CONST.INPROCESS:
            opttask=deepcopy(self.__opttaskindex)
            status=True
            while len(self.__opttaskindex)>0 and status:
                plan,taskno=self.__iter_task(self.__res,opttask)
                if plan==None:
                    status=False
                else:
                    self.__res=plan
                    opttask.remove(taskno)
        return None
    def __cal_vals(self):
        time_usage,power_usage=self.__eval_plan_time_power(self.__res)
        time_efficiency=(sum(time_usage)+sum(self.__PKGTime))/(sum(self.__TTime)+sum(self.__PKGTime))*100
        power_efficiency=(sum(power_usage)+sum(self.__PKGPower))/(sum(self.__TPower)+sum(self.__PKGPower))*100
        res=self.__res[0][:-1]+self.__res[1][:-1]+self.__res[2][:-1]
        revenues=[self.__task.loc[no,'revenue'] for _,no in res]
        objVal=sum(revenues)
        return objVal,time_efficiency,power_efficiency
    def __format_number(self,number):
        number=float(number)
        format_str='{:.'+str(self.__decimal)+'f}'
        formatted_number=format_str.format(number)
        while formatted_number.endswith('0'):
            formatted_number=formatted_number[:-1]
        if formatted_number.endswith('.'):
            formatted_number=formatted_number[:-1]
        return formatted_number
    def _print_status(self,n=None):
        objVal,timeEff,powerEff=self.__cal_vals()
        objVal,timeEff,powerEff=[self.__format_number(number) for number in [objVal,timeEff,powerEff]]
        self.__Status=CONST.SUCCESS
        if n==None:
            print('Feasible objective value '+objVal+' found with '+timeEff+'% time-efficiency and '+powerEff+'% power-efficiency')
        else:
            print('Feasible objective value '+objVal+f' found for Occasion {n} with '+timeEff+'% time-efficiency and '+powerEff+'% power-efficiency')
        return None
    def iter_tags(self):
        n=1
        btimeEff=0.
        bobjVal=0.
        bres=None
        errors=[False,False,False,False,False,False]
        errormessages=['','','','','','']
        for days in ((0,1),(0,2),(1,2)):
            try:
                self._gen_init_plan(n*2-1)
                self._proc_tag(days)
                self._proc_cont()
                self._proc_req()
                self._proc_opt()
                self._add_tag(days)
                self._print_status(n)
                objVal1,timeEff1,_=self.__cal_vals()
                resbackup=deepcopy(self.__res)
            except MIPError as e:
                objVal1,timeEff1=0.,0.
                errors[(n-1)*2]=True
                errormessages[(n-1)*2]=e.message.lower()
                print(f'No feasible solution found for Occasion {n*2-1} because {e.message.lower()}')
            try:
                self._gen_init_plan(n*2)
                self._proc_tag(days)
                self._proc_req()
                self._proc_cont()
                self._proc_opt()
                self._add_tag(days)
                self._print_status(n)
                objVal2,timeEff2,_=self.__cal_vals()
            except MIPError as e:
                objVal2,timeEff2=0.,0.
                errors[(n-1)*2+1]=True
                errormessages[(n-1)*2+1]=e.message.lower()
                print(f'No feasible solution found for Occasion {n*2} because {e.message.lower()}')
            if objVal1>objVal2:
                self.__res=resbackup
                objVal=objVal1
                timeEff=timeEff1
            elif objVal1==objVal2 and timeEff1>timeEff2:
                self.__res=resbackup
                objVal=objVal1
                timeEff=timeEff1
            else:
                objVal=objVal2
                timeEff=timeEff2
            if objVal>bobjVal:
                bres=deepcopy(self.__res)
                btimeEff=timeEff
                bobjVal=objVal
            elif objVal==bobjVal and timeEff>btimeEff:
                bres=deepcopy(self.__res)
                btimeEff=timeEff
                bobjVal=objVal
            n+=1
        if errors[0] and errors[1] and errors[2] and errors[3] and errors[4] and errors[5]:
            errormessages=list(set(errormessages))
            if len(errormessages)==1:
                raise MIPError(f'No feasible solution found because {errormessages[0]}')
            elif len(errormessages)==2:
                raise MIPError(f'No feasible solution found because {errormessages[0]} and {errormessages[0]}')
            else:
                raise MIPError('No feasible solution found')
        print(f'Comparing objective values among {(n-1)*2-1} occasions')
        self.__res=bres
        self._print_status()
        return None
    def cal_route(self):
        self.__res=[self.__res[0][1:],self.__res[1][1:],self.__res[2][1:]]
        pointdf=self.__pointdf.copy()
        pointdf.reset_index(inplace=True)
        pointdf.set_index('No',inplace=True)
        taskdf=self.__task.copy()
        taskdf.reset_index(inplace=True)
        taskdf.set_index('No',inplace=True)
        plan=self.__res.copy()
        n=1
        self.__plandf=[]
        curpt=self.__opoint[0][0]
        for pli in range(3):
            no,action,location,coordinate,time,power,player1,player2,completed=[],[],[],[],[],[],[],[],[]
            if list(self.__opoint[0]) in plan[0]:
                plan[0].remove(list(self.__opoint[0]))
            for i,k in plan[pli]:
                if i==curpt or ((i in (self.__opoint[i][0] for i in range(4))) and (curpt in (self.__opoint[i][0] for i in range(4)))):
                    no.append(n)
                    pti=pointdf.loc[i,'index']
                    ptcur=pointdf.loc[curpt,'index']
                    action.append(taskdf.loc[k,'name'])
                    location.append('探测起点' if pointdf.loc[i,'index'] in ['探测起点1','探测起点2','探测起点3','探测起点4'] else pointdf.loc[i,'index'])
                    coordinate.append(f"({self.__format_number(pointdf.loc[i,'X'])},{self.__format_number(pointdf.loc[i,'Y'])})")
                    time.append(float(self.__format_number(taskdf.loc[k,'time'])))
                    power.append(float(self.__format_number(taskdf.loc[k,'power'])))
                    player1.append('√')
                    player2.append('√')
                    completed.append(np.nan)
                    curpt=i
                    n+=1
                elif i!=curpt:
                    pti=pointdf.loc[i,'index']
                    ptcur=pointdf.loc[curpt,'index']
                    if pti!=ptcur:
                        no.append(n)
                        action.append(f"Travel from ({self.__format_number(pointdf.loc[curpt,'X'])},{self.__format_number(pointdf.loc[curpt,'X'])}) to ({self.__format_number(pointdf.loc[i,'X'])},{self.__format_number(pointdf.loc[i,'Y'])})")
                        ttime=self.__tmatrix.loc[pointdf.loc[curpt,'index']][pointdf.loc[i,'index']]
                        time.append(float(self.__format_number(ttime)))
                        tpower=self.__pmatrix.loc[pointdf.loc[curpt,'index']][pointdf.loc[i,'index']]
                        power.append(float(self.__format_number(tpower)))
                        location.append(f"{'探测起点' if pointdf.loc[curpt,'index'] in ['探测起点1','探测起点2','探测起点3','探测起点4'] else pointdf.loc[curpt,'index']}→{'探测起点' if pointdf.loc[i,'index'] in ['探测起点1','探测起点2','探测起点3','探测起点4'] else pointdf.loc[i,'index']}")
                        coordinate.append(f"({self.__format_number(pointdf.loc[curpt,'X'])},{self.__format_number(pointdf.loc[curpt,'Y'])})→({self.__format_number(pointdf.loc[i,'X'])},{self.__format_number(pointdf.loc[i,'Y'])})")
                        player1.append('√')
                        player2.append('√')
                        completed.append(np.nan)
                        n+=1
                    no.append(n)
                    action.append(taskdf.loc[k,'name'])
                    location.append('探测起点' if pointdf.loc[i,'index'] in ['探测起点1','探测起点2','探测起点3','探测起点4'] else pointdf.loc[i,'index'])
                    coordinate.append(f"({self.__format_number(pointdf.loc[i,'X'])},{self.__format_number(pointdf.loc[i,'Y'])})")
                    time.append(float(self.__format_number(taskdf.loc[k,'time'])))
                    power.append(float(self.__format_number(taskdf.loc[k,'power'])))
                    player1.append('√')
                    player2.append('√')
                    completed.append(np.nan)
                    curpt=i
                    n+=1
            no.pop()
            action.pop()
            location.pop()
            coordinate.pop()
            time.pop()
            power.pop()
            player1.pop()
            player2.pop()
            completed.pop()
            self.__plandf.append(pd.DataFrame({'No':no,'action':action,'location':location,'coordinate':coordinate,'time':time,'power':power,'1':player1,'2':player2,'completed':completed}))
        return None
    def __gen_packagedf(self,package,tag):
        pak=self.__package[package['tag']==tag]
        no=[i for i in range(pak.shape[0])]
        location=f"({self.__format_number(self.__pointdf.loc['探测起点1','X'])},{self.__format_number(self.__pointdf.loc['探测起点1','X'])})"
        coord=[location for i in range(pak.shape[0])]
        loc=['探测起点' for i in range(pak.shape[0])]
        player=['√' for i in range(pak.shape[0])]
        completed=[np.nan for i in range(pak.shape[0])]
        df=pd.DataFrame({'No':no,'action':pak.loc[:,'name'].values,'location':loc,'coordinate':coord,'time':pak.loc[:,'time'].values,'power':pak.loc[:,'power'].values,'1':player,'2':player,'completed':completed})
        return df
    def add_package(self):
        plan=self.__plandf.copy()
        isindf12=[False,False,False]
        isindf23=[False,False,False]
        df=[False,False,False]
        df0=pd.DataFrame({'No':[np.nan],'action':['Begin of Day1'],'location':[np.nan],'coordinate':[np.nan],'time':[np.nan],'power':[np.nan],'1':[np.nan],'2':[np.nan],'completed':[np.nan]})
        for i in range(3):
            isindf12[i]=self.__plandf[i].isin(self.__task[self.__task['tag']=='12s']['name'].values).any().any()
            isindf23[i]=self.__plandf[i].isin(self.__task[self.__task['tag']=='23s']['name'].values).any().any()
        pakdf=[[[self.__gen_packagedf(self.__package,'D1ss'),self.__gen_packagedf(self.__package,'D1se')],[self.__gen_packagedf(self.__package,'D1es'),self.__gen_packagedf(self.__package,'D1ee')]],
                [[self.__gen_packagedf(self.__package,'D2ss'),self.__gen_packagedf(self.__package,'D2se')],[self.__gen_packagedf(self.__package,'D2es'),self.__gen_packagedf(self.__package,'D2ee')]],
                [[self.__gen_packagedf(self.__package,'D3ss'),self.__gen_packagedf(self.__package,'D3se')],[self.__gen_packagedf(self.__package,'D3es'),self.__gen_packagedf(self.__package,'D3ee')]]]
        for i in range(3):
            if isindf12[i]:
                if self.__plandf[i].shape[0]==2:
                    temp=pd.concat([self.__plandf[i].iloc[:1,:],pakdf[i][0][1],pakdf[i][1][0],self.__plandf[i].iloc[-1:,:]])
                elif self.__plandf[i].shape[0]>2:
                    temp=pd.concat([self.__plandf[i].iloc[:1,:],pakdf[i][0][1],self.__plandf[i].iloc[1:-1,:],pakdf[i][1][0],self.__plandf[i].iloc[-1:,:]])
                time=sum(temp['time'])
                if time<=self.__12gap:
                    adddf=pd.DataFrame({'No':[np.nan],'action':[f'Wait for {self.__format_number(self.__12gap-time)}s'],'location':['探测起点'],'coordinate':[f"({self.__format_number(self.__pointdf.loc['探测起点1','X'])},{self.__format_number(self.__pointdf.loc['探测起点1','X'])})"],'time':[float(self.__format_number(self.__12gap-time))],'power':[0],'1':['√'],'2':['√']})
                    if self.__plandf[i].shape[0]==2:
                        plan[i]=pd.concat([pakdf[i][0][0],self.__plandf[i].iloc[:1,:],pakdf[i][0][1],pakdf[i][1][0],adddf,self.__plandf[i].iloc[-1:,:],pakdf[i][1][1]])
                    elif self.__plandf[i].shape[0]>2:
                        plan[i]=pd.concat([pakdf[i][0][0],self.__plandf[i].iloc[:1,:],pakdf[i][0][1],self.__plandf[i].iloc[1:-1,:],pakdf[i][1][0],adddf,self.__plandf[i].iloc[-1:,:],pakdf[i][1][1]])
                else:
                    if self.__plandf[i].shape[0]==2:
                        plan[i]=pd.concat([pakdf[i][0][0],self.__plandf[i].iloc[:1,:],pakdf[i][0][1],pakdf[i][1][0],self.__plandf[i].iloc[-1:,:],pakdf[i][1][1]])
                    elif self.__plandf[i].shape[0]>2:
                        plan[i]=pd.concat([pakdf[i][0][0],self.__plandf[i].iloc[:1,:],pakdf[i][0][1],self.__plandf[i].iloc[1:-1,:],pakdf[i][1][0],self.__plandf[i].iloc[-1:,:],pakdf[i][1][1]])
                time='sum: {}'.format(self.__format_number(sum(plan[i]['time'])))
                power='sum: {}'.format(self.__format_number(sum(plan[i]['power'])))
                df[i]=pd.DataFrame({'No':[np.nan],'action':['xxx'],'location':[np.nan],'coordinate':[np.nan],'time':[time],'power':[power],'1':[np.nan],'2':[np.nan],'completed':[np.nan]})
            elif isindf23[i]:
                if self.__plandf[i].shape[0]==2:
                    temp=pd.concat([self.__plandf[i].iloc[:1,:],pakdf[i][0][1],pakdf[i][1][0],self.__plandf[i].iloc[-1:,:]])
                elif self.__plandf[i].shape[0]>2:
                    temp=pd.concat([self.__plandf[i].iloc[:1,:],pakdf[i][0][1],self.__plandf[i].iloc[1:-1,:],pakdf[i][1][0],self.__plandf[i].iloc[-1:,:]])
                time=sum(temp['time'])
                if time<=self.__23gap:
                    adddf=pd.DataFrame({'No':[np.nan],'action':[f'Wait for {self.__format_number(self.__23gap-time)}s'],'location':['探测起点'],'coordinate':[f"({self.__format_number(self.__pointdf.loc['探测起点1','X'])},{self.__format_number(self.__pointdf.loc['探测起点1','X'])})"],'time':[float(self.__format_number(self.__23gap-time))],'power':[0],'1':['√'],'2':['√']})
                    if self.__plandf[i].shape[0]==2:
                        plan[i]=pd.concat([pakdf[i][0][0],self.__plandf[i].iloc[:1,:],pakdf[i][0][1],pakdf[i][1][0],adddf,self.__plandf[i].iloc[-1:,:],pakdf[i][1][1]])
                    elif self.__plandf[i].shape[0]>2:
                        plan[i]=pd.concat([pakdf[i][0][0],self.__plandf[i].iloc[:1,:],pakdf[i][0][1],self.__plandf[i].iloc[1:-1,:],pakdf[i][1][0],adddf,self.__plandf[i].iloc[-1:,:],pakdf[i][1][1]])
                else:
                    if self.__plandf[i].shape[0]==2:
                        plan[i]=pd.concat([pakdf[i][0][0],self.__plandf[i].iloc[:1,:],pakdf[i][0][1],pakdf[i][1][0],self.__plandf[i].iloc[-1:,:],pakdf[i][1][1]])
                    elif self.__plandf[i].shape[0]>2:
                        plan[i]=pd.concat([pakdf[i][0][0],self.__plandf[i].iloc[:1,:],pakdf[i][0][1],self.__plandf[i].iloc[1:-1,:],pakdf[i][1][0],self.__plandf[i].iloc[-1:,:],pakdf[i][1][1]])
                time='sum: {}'.format(self.__format_number(sum(plan[i]['time'])))
                power='sum: {}'.format(self.__format_number(sum(plan[i]['power'])))
                df[i]=pd.DataFrame({'No':[np.nan],'action':['xxx'],'location':[np.nan],'coordinate':[np.nan],'time':[time],'power':[power],'1':[np.nan],'2':[np.nan],'completed':[np.nan]})
            else:
                pakdf[i][0]=pd.concat(pakdf[i][0])
                pakdf[i][1]=pd.concat(pakdf[i][1])
                plan[i]=pd.concat([pakdf[i][0],self.__plandf[i],pakdf[i][1]])
                time='sum: {}'.format(self.__format_number(sum(plan[i]['time'])))
                power='sum: {}'.format(self.__format_number(sum(plan[i]['power'])))
                df[i]=pd.DataFrame({'No':[np.nan],'action':['xxx'],'location':[np.nan],'coordinate':[np.nan],'time':[time],'power':[power],'1':[np.nan],'2':[np.nan],'completed':[np.nan]})
        df[0]['action']='Break between Day1 and Day2'
        df[1]['action']='Break between Day2 and Day3'
        df[2]['action']='End of Day3'
        self.__plandf=pd.concat([df0,plan[0],df[0],plan[1],df[1],plan[2],df[2]])
        self.__plandf.reset_index(drop=True,inplace=True)
        index0=self.__plandf[self.__plandf['action']=='Begin of Day1'].index[0]+2
        index1=self.__plandf[self.__plandf['action']=='Break between Day1 and Day2'].index[0]+2
        index2=self.__plandf[self.__plandf['action']=='Break between Day2 and Day3'].index[0]+2
        index3=self.__plandf[self.__plandf['action']=='End of Day3'].index[0]+2
        self.__voidindex=[int(index0),int(index1),int(index2),int(index3)]
        no=[np.nan]+list(range(1,index1-2))+[np.nan]+list(range(index1-2,index2-3))+[np.nan]+list(range(index2-3,index3-4))+[np.nan]
        self.__plandf['No']=no
        return None
    def write_excel(self):
        with pd.ExcelWriter(self.__outputpath) as writer:
            self.__plandf.to_excel(writer,index=False)
            worksheet=writer.sheets[list(writer.sheets.keys())[0]]
            for column_cells in worksheet.columns:
                lengthlist=[]
                for cell in column_cells:
                    length=0
                    for char in str(cell.value):
                        if char>=u'\u4e00' and char<=u'\u9fff':
                            length+=2
                        else:
                            length+=1
                    lengthlist.append(length)
                length=max(max(lengthlist),5)
                worksheet.column_dimensions[column_cells[0].column_letter].width=length+2
            for row in self.__voidindex:
                for cell in worksheet[row]:
                    cell.fill=PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
        timecost=self.__format_number(time.time()-self.__starttime)
        print('Process finished in '+timecost+' seconds')
        return None
    def schedule_frame(self):
        return self.__plandf.copy()
    def run(self):
        self.test_IO()
        self.read_info()
        self.read_task()
        self.read_package()
        self.read_point()
        self.gen_void_point()
        self.check_remote()
        self.divide_task()
        self.drop_O()
        self.iter_tags()
        self.cal_route()
        self.add_package()
        if self.__write_output:
            self.write_excel()
        return None
class heuristic_revisional(object):
    def __init__(self,workPath=None,decimal=5,timePathRevised=None,powerPathRevised=None,
                 infoPath='info.xlsx',taskPath='task.xlsx',packPath='package.xlsx',
                 pointPath='point.xlsx',distancePath='distance.xlsx',timePath='time.xlsx',
                 powerPath='power.xlsx',schedulePath='schedule.xlsx',outputPath='schedule-revised.xlsx',
                 dataFrames=None,scheduleFrame=None,writeOutput=True):
        print('Process initiated')
        self.__starttime=time.time()
        self.__decimal=decimal
        if workPath==None:
            self.__infoPath,self.__taskPath,self.__pakpath=infoPath,taskPath,packPath
            self.__pointpath,self.__distancepath,self.__timepath,self.__powerpath=pointPath,distancePath,timePath,powerPath
        else:
            self.__infoPath,self.__taskPath,self.__pakpath=os.path.join(workPath,infoPath),os.path.join(workPath,taskPath),os.path.join(workPath,packPath)
            self.__pointpath,self.__distancepath,self.__timepath,self.__powerpath=os.path.join(workPath,pointPath),os.path.join(workPath,distancePath),os.path.join(workPath,timePath),os.path.join(workPath,powerPath)
        self.__timepathrevised,self.__powerpathrevised=timePathRevised,powerPathRevised
        self.__schedulepath=schedulePath
        self.__outputpath=outputPath
        self.__dataframes=dataFrames or {}
        self.__scheduleframe=scheduleFrame
        self.__write_output=writeOutput
        self.__Status=CONST.CREATED
    def __trans_time(self,time):
        newtime=time*2
        return newtime
    def __trans_power(self,power):
        newpower=power*1.5
        return newpower
    def test_IO(self):
        if self.__dataframes:
            missing=[key for key in ['info','task','package','point','distance','time','power'] if key not in self.__dataframes]
            if missing:
                raise ValueError(f'Missing DataFrame inputs: {missing}')
            if self.__scheduleframe is None:
                raise ValueError('Missing scheduleFrame input')
            return None
        FileNotFound=[]
        if not os.path.exists(self.__infoPath):
            FileNotFound.append(self.__infoPath)
        if not os.path.exists(self.__taskPath):
            FileNotFound.append(self.__taskPath)
        if not os.path.exists(self.__pakpath):
            FileNotFound.append(self.__pakpath)
        if not os.path.exists(self.__pointpath):
            FileNotFound.append(self.__pointpath)
        if not os.path.exists(self.__distancepath):
            FileNotFound.append(self.__distancepath)
        if not os.path.exists(self.__timepath):
            FileNotFound.append(self.__timepath)
        if not os.path.exists(self.__powerpath):
            FileNotFound.append(self.__powerpath)
        if not os.path.exists(self.__schedulepath):
            FileNotFound.append(self.__schedulepath)
        if self.__timepathrevised != None:
            if not os.path.exists(self.__timepathrevised):
                FileNotFound.append(self.__timepathrevised)
        if self.__powerpathrevised != None:
            if not os.path.exists(self.__powerpathrevised):
                FileNotFound.append(self.__powerpathrevised)
        if FileNotFound!=[]:
            raise FileNotFoundError('No such file: {}'.format(' '.join(FileNotFound)))
        notPermitted=False
        if self.__write_output and os.path.exists(self.__outputpath):
            try:
                pd.read_excel(self.__outputpath)
            except PermissionError:
                notPermitted=True
        if notPermitted:
            raise PermissionError(f'Permission denied: {self.__outputpath}')
        return None
    def read_info(self):
        info=self.__dataframes.get('info')
        info=info.copy() if info is not None else pd.read_excel(self.__infoPath)
        self.__MDistance=info['max-distance'][0]
        self.__TTime=list(map(float,info['total-time/day'][0].split(';')))
        self.__TPower=list(map(float,info['total-power/day'][0].split(';')))
        self.__TTimebackup=deepcopy(self.__TTime)
        self.__TPowerbackup=deepcopy(self.__TPower)
        self.__Mincontinuous=info['min-continuous'][0]
        self.__12gap=info['12-gap'][0]
        self.__23gap=info['23-gap'][0]
        return None
    def __read_matrix(self,path,key=None):
        matrix=self.__dataframes.get(key) if key is not None else None
        if matrix is not None:
            matrix=matrix.copy()
        else:
            matrix=pd.read_excel(path)
            matrix.set_index(matrix.columns[0],inplace=True)
            matrix.index.rename(None,inplace=True)
        return matrix
    def read_task(self):
        task=self.__dataframes.get('task')
        self.__task=task.copy() if task is not None else pd.read_excel(self.__taskPath)
        self.__task['No']=range(self.__task.shape[0])
        self.__dmatrix=self.__read_matrix(self.__distancepath,'distance')
        self.__dmatrix.replace(np.inf,self.__MDistance*3,inplace=True)
        np.fill_diagonal(self.__dmatrix.values,0)
        self.__tmatrix=self.__read_matrix(self.__timepath,'time')
        self.__tmatrix.replace(np.inf,max(self.__TTime),inplace=True)
        np.fill_diagonal(self.__tmatrix.values,0)
        self.__pmatrix=self.__read_matrix(self.__powerpath,'power')
        self.__pmatrix.replace(np.inf,max(self.__TPower),inplace=True)
        np.fill_diagonal(self.__pmatrix.values,0)
        return None
    def read_package(self):
        package=self.__dataframes.get('package')
        self.__package=package.copy() if package is not None else pd.read_excel(self.__pakpath)
        time=[sum(self.__package[self.__package['tag']==tag]['time']) for tag in ['D1ss','D1se','D1es','D1ee','D2ss','D2se','D2es','D2ee','D3ss','D3se','D3es','D3ee']]
        power=[sum(self.__package[self.__package['tag']==tag]['power']) for tag in ['D1ss','D1se','D1es','D1ee','D2ss','D2se','D2es','D2ee','D3ss','D3se','D3es','D3ee']]
        self.__PKGTime=[time[0]+time[1]+time[2]+time[3],time[4]+time[5]+time[6]+time[7],time[8]+time[9]+time[10]+time[11]]
        self.__PKGPower=[power[0]+power[1]+power[2]+power[3],power[4]+power[5]+power[6]+power[7],power[8]+power[9]+power[10]+power[11]]
        self.__TTime=[self.__TTime[i]-self.__PKGTime[i] for i in range(3)]
        self.__TPower=[self.__TPower[i]-self.__PKGPower[i] for i in range(3)]
        return None
    def read_point(self):
        point=self.__dataframes.get('point')
        self.__pointdf=point.copy() if point is not None else pd.read_excel(self.__pointpath)
        self.__pointdf.set_index(self.__pointdf.columns[0],inplace=True)
        self.__pointdf.index.rename(None,inplace=True)
        return None
    def __add_void_matrix(self,matrix):
        voidrow=[matrix.loc['探测起点',:].values.tolist() for i in range(4)]
        voidrowdf=pd.DataFrame(voidrow,index=['探测起点1','探测起点2','探测起点3','探测起点4'],columns=matrix.columns)
        matrix=pd.concat([voidrowdf,matrix])
        matrix.drop(['探测起点'],inplace=True)
        voidcol=np.array([matrix.loc[:,'探测起点'].values.tolist() for i in range(4)]).T
        voidcoldf=pd.DataFrame(voidcol,columns=['探测起点1','探测起点2','探测起点3','探测起点4'],index=matrix.index)
        matrix=pd.concat([voidcoldf,matrix],axis=1)
        matrix.drop(['探测起点'],axis=1,inplace=True)
        for i,j in product(['探测起点1','探测起点2','探测起点3','探测起点4'],['探测起点1','探测起点2','探测起点3','探测起点4']):
            if i!=j:
                matrix.loc[i,j]=0
        return matrix
    def gen_void_point(self):
        voidpoint=pd.DataFrame({'X':[self.__pointdf.loc['探测起点','X'],self.__pointdf.loc['探测起点','X'],self.__pointdf.loc['探测起点','X'],self.__pointdf.loc['探测起点','X']],
                                'Y':[self.__pointdf.loc['探测起点','Y'],self.__pointdf.loc['探测起点','Y'],self.__pointdf.loc['探测起点','Y'],self.__pointdf.loc['探测起点','Y']],
                                '备注':['虚拟原点：第一天出发','虚拟原点：第一天返回第二天出发','虚拟原点：第二天返回第三天出发','虚拟原点：第三天返回']},
                               index=['探测起点1','探测起点2','探测起点3','探测起点4'])
        self.__pointdf=pd.concat([voidpoint,self.__pointdf])
        self.__pointdf.drop(['探测起点'],inplace=True)
        self.__pointdf['No']=range(self.__pointdf.shape[0])
        self.__dmatrix=self.__add_void_matrix(self.__dmatrix)
        self.__tmatrix=self.__add_void_matrix(self.__tmatrix)
        self.__pmatrix=self.__add_void_matrix(self.__pmatrix)
        self.__task['location']=self.__task['location'].replace('探测起点','探测起点1,探测起点2,探测起点3,探测起点4')
        self.__task['location']=self.__task['location'].fillna(','.join(self.__pointdf.index.values))
        location=self.__task['location'].values
        pool=[]
        for i in location:
            if pd.isna(i):
                pool.append(np.nan)
            else:
                split=list(i.split(','))
                pool.append(split)
        self.__task['location']=pd.Series(pool)
        voidtaskdf=pd.DataFrame({'No':[self.__task.shape[0],self.__task.shape[0]+1,self.__task.shape[0]+2,self.__task.shape[0]+3],
                                 'name':['void1','void2','void3','void4'],
                                 'revenue':[0,0,0,0],
                                 'location':[['探测起点1'],['探测起点2'],['探测起点3'],['探测起点4']],
                                 'day':[np.nan,np.nan,np.nan,np.nan],
                                 'time':[0,0,0,0],
                                 'power':[0,0,0,0],
                                 'required':[False,False,False,False],
                                 'continuous':[False,False,False,False],
                                 'remote':[False,False,False,False],
                                 'exceptO':[False,False,False,False],
                                 'tag':[np.nan,np.nan,np.nan,np.nan]},
                                index=[self.__task.shape[0],self.__task.shape[0]+1,self.__task.shape[0]+2,self.__task.shape[0]+3])
        self.__task=pd.concat([self.__task,voidtaskdf])
        self.__opoint=((self.__pointdf.loc['探测起点1','No'],self.__task[self.__task['name']=='void1'].index[0]),
                       (self.__pointdf.loc['探测起点2','No'],self.__task[self.__task['name']=='void2'].index[0]),
                       (self.__pointdf.loc['探测起点3','No'],self.__task[self.__task['name']=='void3'].index[0]),
                       (self.__pointdf.loc['探测起点4','No'],self.__task[self.__task['name']=='void4'].index[0]))
        self.__pointdfbackup=self.__pointdf.copy()
        self.__pointdfbackup.reset_index(inplace=True)
        self.__pointdfbackup.set_index('No',inplace=True)
        return None
    def __cartesian_to_polar(self,x,y,O):
        r=np.sqrt((x-O[0])**2+(y-O[1])**2)
        theta=np.arctan2(y-O[1],x-O[0])
        return r,theta
    def __polar_to_cartesian(self,r,theta,O):
        x=r*np.cos(theta)+O[0]
        y=r*np.sin(theta)+O[1]
        return x,y
    def __cal_distance(self,point1,point2):
        dis=((point1[0]-point2[0])**2+(point1[1]-point2[1])**2)**0.5
        return dis
    def __from_distance(self,distance):
        pt=(distance,0)
        return pt
    def check_remote(self):
        remoteindex=self.__task[self.__task['remote']==True].index.to_list()
        points=[]
        for index in remoteindex:
            pts=[]
            for point in self.__task.loc[index,'location']:
                if self.__dmatrix.loc['探测起点1',point]>=self.__MDistance:
                    points.append(point)
                    pts.append(point)
            self.__task.at[index,'location']=pts
        if len(points)==0:
            start=0
            path='new-point.xlsx'
            sortedseries=self.__dmatrix.sort_values(by=['探测起点1'],ascending=False).loc['探测起点1',:]
            while sortedseries[start]>=self.__MDistance:
                start+=1
            pts=self.__dmatrix.sort_values(by=['探测起点1'],ascending=False).iloc[start:start+5,:].index.values
            r,_=self.__cartesian_to_polar(self.__from_distance(self.__MDistance)[0],self.__from_distance(self.__MDistance)[1],(0,0))
            pts=[self.__cartesian_to_polar(self.__pointdf.loc[pt,'X'],self.__pointdf.loc[pt,'Y'],(self.__pointdf.loc['探测起点1','X'],self.__pointdf.loc['探测起点1','Y'])) for pt in pts]
            pts=[self.__polar_to_cartesian(r,pt[1],(self.__pointdf.loc['探测起点1','X'],self.__pointdf.loc['探测起点1','Y'])) for pt in pts]
            newpt=pd.DataFrame(pts,columns=['X','Y'],index=pd.Index(['新最远探测点1','新最远探测点2','新最远探测点3','新最远探测点4','新最远探测点5'],name='name'))
            raise MaxDistanceError(f'No point meets the max distance requirement. Recommended points have been generated in {path}',newpt,path)
        return None
    def divide_task(self):
        self.__reqtaskindex=self.__task[self.__task['required']==True].index.to_list()
        self.__opttaskindex=self.__task[self.__task['required']==False].index.to_list()
        self.__daytaskindex=[self.__task[self.__task['day']==1].index.to_list(),
                             self.__task[self.__task['day']==2].index.to_list(),
                             self.__task[self.__task['day']==3].index.to_list(),
                             self.__task[self.__task['day']=='1,2'].index.to_list(),
                             self.__task[self.__task['day']=='2,3'].index.to_list(),
                             self.__task[self.__task['day']=='1,3'].index.to_list()]
        self.__remtaskindex=self.__task[self.__task['remote']==True].index.to_list()
        self.__tagtaskindex=[self.__task[self.__task['tag']=='12s'].index.to_list(),
                             self.__task[self.__task['tag']=='12e'].index.to_list(),
                             self.__task[self.__task['tag']=='23s'].index.to_list(),
                             self.__task[self.__task['tag']=='23e'].index.to_list()]
        self.__contaskindex=self.__task[self.__task['continuous']==True].index.to_list()
        tagtask=[]
        for i in range(len(self.__tagtaskindex)):
            tagtask.append(*self.__tagtaskindex[i])
        for index in self.__reqtaskindex.copy():
            if index in self.__contaskindex:
                self.__reqtaskindex.remove(index)
            if index in tagtask:
                self.__reqtaskindex.remove(index)
        for index in self.__opttaskindex.copy():
            if index in self.__contaskindex:
                self.__opttaskindex.remove(index)
            if index in tagtask:
                self.__opttaskindex.remove(index)
            if index in [self.__opoint[0][1],self.__opoint[1][1],self.__opoint[2][1],self.__opoint[3][1]]:
                self.__opttaskindex.remove(index)
        self.__noOtaskindex=self.__task[self.__task['exceptO']==True].index.to_list()
        return None
    def drop_O(self):
        for i in self.__noOtaskindex:
            self.__task.loc[i,'location'].remove('探测起点1')
            self.__task.loc[i,'location'].remove('探测起点2')
            self.__task.loc[i,'location'].remove('探测起点3')
            self.__task.loc[i,'location'].remove('探测起点4')
        return None
    def update_matrix(self):
        if self.__timepathrevised != None:
            self.__tmatrix=self.__read_matrix(self.__timepathrevised)
            self.__tmatrix.replace(np.inf,max(self.__TTime),inplace=True)
            np.fill_diagonal(self.__tmatrix.values,0)
            self.__tmatrix=self.__add_void_matrix(self.__tmatrix)
        else:
            self.__tmatrix=self.__tmatrix.map(self.__trans_time)
        if self.__powerpathrevised != None:
            self.__pmatrix=self.__read_matrix(self.__powerpathrevised)
            self.__pmatrix.replace(np.inf,max(self.__TPower),inplace=True)
            np.fill_diagonal(self.__pmatrix.values,0)
            self.__pmatrix=self.__add_void_matrix(self.__pmatrix)
        else:
            self.__pmatrix=self.__pmatrix.map(self.__trans_power)
        return None
    def read_schedule(self):
        schedule=self.__scheduleframe.copy() if self.__scheduleframe is not None else pd.read_excel(self.__schedulepath)
        index=schedule[schedule['No'].isna()].index.to_list()
        schedule=[schedule.iloc[index[0]+1:index[1],:],schedule.iloc[index[1]+1:index[2],:],schedule.iloc[index[2]+1:index[3],:]]
        self.__completed_schedule=[schedule[0][schedule[0]['completed'].notna()],schedule[1][schedule[1]['completed'].notna()],schedule[2][schedule[2]['completed'].notna()]]
        self.__current_state=[None,[[self.__opoint[0],self.__opoint[1]],[self.__opoint[1],self.__opoint[2]],[self.__opoint[2],self.__opoint[3]]],('D1ss','D1se','D1es','D1ee','D2ss','D2se','D2es','D2ee','D3ss','D3se','D3es','D3ee')]
        return None
    def __find_point(self,X,Y):
        eps=1/(10**self.__decimal)/2
        ptno=self.__pointdf[(self.__pointdf['X']>=X-eps) & (self.__pointdf['X']<X+eps) & (self.__pointdf['Y']>=Y-eps) & (self.__pointdf['Y']<Y+eps)]['No'].values[0]
        return ptno
    def update_task(self):
        day=0
        self.__completed_tag=[None,None,None,None]
        for schedule in self.__completed_schedule:
            for _,row in self.__completed_schedule[day].iterrows():
                task=self.__task[self.__task['name']==row['action']]
                if len(task['tag'].values)>0:
                    if task['tag'].values[0]=='12s':
                        self.__completed_tag[0]=day
                        self.__tagtaskindex[0]=None
                        self.__current_state[0]='12s'
                        self.__current_state[1][day][0]=(self.__opoint[day][0],self.__current_state[1][day][0][1])
                    elif task['tag'].values[0]=='12e':
                        self.__completed_tag[1]=day
                        self.__tagtaskindex[1]=None
                        self.__current_state[0]='12e'
                        self.__current_state[1][day][0]=(self.__opoint[day][0],self.__current_state[1][day][0][1])
                    elif task['tag'].values[0]=='23s':
                        self.__completed_tag[2]=day
                        self.__tagtaskindex[2]=None
                        self.__current_state[0]='23s'
                        self.__current_state[1][day][0]=(self.__opoint[day][0],self.__current_state[1][day][0][1])
                    elif task['tag'].values[0]=='23e':
                        self.__completed_tag[3]=day
                        self.__tagtaskindex[3]=None
                        self.__current_state[0]='23e'
                        self.__current_state[1][day][0]=(self.__opoint[day][0],self.__current_state[1][day][0][1])
                    else:
                        if row['coordinate'].count('→')==0:
                            X,Y=list(row['coordinate'].split(','))
                        elif row['coordinate'].count('→')==1:
                            X,Y=list(row['coordinate'].split(',')[1].split(','))
                        X,Y=float(X[1:]),float(Y[:-1])
                        self.__current_state[1][day][0]=(self.__find_point(X,Y),self.__current_state[1][day][0][1])
                        index=task.index.to_list()[0]
                        if index in self.__reqtaskindex:
                            self.__reqtaskindex.remove(index)
                        if index in self.__opttaskindex:
                            self.__opttaskindex.remove(index)
                        if index in self.__opttaskindex:
                            self.__opttaskindex.remove(index)
                        for i in range(len(self.__daytaskindex)):
                            if index in self.__daytaskindex[i]:
                                self.__daytaskindex[i].remove(index)
                        if index in self.__remtaskindex:
                            self.__remtaskindex.remove(index)
                        if index in self.__contaskindex:
                            self.__contaskindex.remove(index)
                            self.__Mincontinuous-=1
                self.__task.drop(task.index,inplace=True)
            day+=1
        return None
    def update_package(self):
        day=0
        for schedule in self.__completed_schedule:
            for _,row in self.__completed_schedule[day].iterrows():
                task=self.__package[self.__package['name']==row['action']]
                task=task[task['tag'].str[1:2]==str(day+1)]
                if len(task['tag'].values)>0:
                    self.__package.drop(task.index,inplace=True)
            day+=1
        time=[sum(self.__package[self.__package['tag']==tag]['time']) for tag in ['D1ss','D1se','D1es','D1ee','D2ss','D2se','D2es','D2ee','D3ss','D3se','D3es','D3ee']]
        power=[sum(self.__package[self.__package['tag']==tag]['power']) for tag in ['D1ss','D1se','D1es','D1ee','D2ss','D2se','D2es','D2ee','D3ss','D3se','D3es','D3ee']]
        self.__PKGTime=[time[0]+time[1]+time[2]+time[3],time[4]+time[5]+time[6]+time[7],time[8]+time[9]+time[10]+time[11]]
        self.__PKGPower=[power[0]+power[1]+power[2]+power[3],power[4]+power[5]+power[6]+power[7],power[8]+power[9]+power[10]+power[11]]
        self.__CompTime=[self.__completed_schedule[0]['time'].sum(),self.__completed_schedule[1]['time'].sum(),self.__completed_schedule[2]['time'].sum()]
        self.__CompPower=[self.__completed_schedule[0]['power'].sum(),self.__completed_schedule[1]['power'].sum(),self.__completed_schedule[2]['power'].sum()]
        self.__TTime=[self.__TTimebackup[i]-self.__PKGTime[i]-self.__CompTime[i] for i in range(3)]
        self.__TPower=[self.__TPowerbackup[i]-self.__PKGPower[i]-self.__CompPower[i] for i in range(3)]
        return None
    def _gen_init_plan(self,n=None):
        if n==None:
            print('Solving feasible plan')
        else:
            print(f'Solving feasible plan for Occasion {n}')
        self.__Status=CONST.INPROCESS
        self.__res=deepcopy(self.__current_state[1])
        return None
    def _proc_tag(self,days=(0,1)):
        if self.__Status==CONST.INPROCESS:
            self.__TTimeRev=deepcopy(self.__TTime)
            self.__TPowerRev=deepcopy(self.__TPower)
            if days[0]==0:
                if self.__current_state[0]==None:
                    self.__TTimeRev[0]=self.__TTimeRev[0]-self.__task.loc[self.__task[self.__task['tag']=='12s'].index[0],'time']-self.__task.loc[self.__task[self.__task['tag']=='12e'].index[0],'time']
                    self.__TPowerRev[0]=self.__TPowerRev[0]-self.__task.loc[self.__task[self.__task['tag']=='12s'].index[0],'power']-self.__task.loc[self.__task[self.__task['tag']=='12e'].index[0],'power']
                if self.__current_state[0]=='12s':
                    self.__TTimeRev[0]=self.__TTimeRev[0]-self.__task.loc[self.__task[self.__task['tag']=='12e'].index[0],'time']
                    self.__TPowerRev[0]=self.__TPowerRev[0]-self.__task.loc[self.__task[self.__task['tag']=='12e'].index[0],'power']
                elif self.__current_state[0] in ('12e','23s','23e'):
                    pass
            elif days[0]==1:
                if self.__current_state[0]==None:
                    self.__TTimeRev[1]=self.__TTimeRev[1]-self.__task.loc[self.__task[self.__task['tag']=='12s'].index[0],'time']-self.__task.loc[self.__task[self.__task['tag']=='12e'].index[0],'time']
                    self.__TPowerRev[1]=self.__TPowerRev[1]-self.__task.loc[self.__task[self.__task['tag']=='12s'].index[0],'power']-self.__task.loc[self.__task[self.__task['tag']=='12e'].index[0],'power']
                if self.__current_state[0]=='12s':
                    self.__TTimeRev[1]=self.__TTimeRev[1]-self.__task.loc[self.__task[self.__task['tag']=='12e'].index[0],'time']
                    self.__TPowerRev[1]=self.__TPowerRev[1]-self.__task.loc[self.__task[self.__task['tag']=='12e'].index[0],'power']
                elif self.__current_state[0] in ('12e','23s','23e'):
                    pass
            if days[1]==1:
                if self.__current_state[0] in (None,'12s','12e'):
                    self.__TTimeRev[1]=self.__TTimeRev[1]-self.__task.loc[self.__task[self.__task['tag']=='23s'].index[0],'time']-self.__task.loc[self.__task[self.__task['tag']=='23e'].index[0],'time']
                    self.__TPowerRev[1]=self.__TPowerRev[1]-self.__task.loc[self.__task[self.__task['tag']=='23s'].index[0],'power']-self.__task.loc[self.__task[self.__task['tag']=='23e'].index[0],'power']
                elif self.__current_state[0]=='23s':
                    self.__TTimeRev[1]=self.__TTimeRev[1]-self.__task.loc[self.__task[self.__task['tag']=='23e'].index[0],'time']
                    self.__TPowerRev[1]=self.__TPowerRev[1]-self.__task.loc[self.__task[self.__task['tag']=='23e'].index[0],'power']
                elif self.__current_state[0]=='23e':
                    self.__TTimeRev[1]=self.__TTimeRev[1]
                    self.__TPowerRev[1]=self.__TPowerRev[1]
            elif days[1]==2:
                if self.__current_state[0] in (None,'12s','12e'):
                    self.__TTimeRev[2]=self.__TTimeRev[2]-self.__task.loc[self.__task[self.__task['tag']=='23s'].index[0],'time']-self.__task.loc[self.__task[self.__task['tag']=='23e'].index[0],'time']
                    self.__TPowerRev[2]=self.__TPowerRev[2]-self.__task.loc[self.__task[self.__task['tag']=='23s'].index[0],'power']-self.__task.loc[self.__task[self.__task['tag']=='23e'].index[0],'power']
                elif self.__current_state[0]=='23s':
                    self.__TTimeRev[2]=self.__TTimeRev[2]-self.__task.loc[self.__task[self.__task['tag']=='23e'].index[0],'time']
                    self.__TPowerRev[2]=self.__TPowerRev[2]-self.__task.loc[self.__task[self.__task['tag']=='23e'].index[0],'power']
                elif self.__current_state[0]=='23e':
                    pass
        return None
    def __check_plan(self,plan,day,node):
        time,power=self.__eval_plan_time_power(plan)
        timeOK=time[0]<=self.__TTimeRev[0] and time[1]<=self.__TTimeRev[1] and time[2]<=self.__TTimeRev[2]
        powerOK=power[0]<=self.__TPowerRev[0] and power[1]<=self.__TPowerRev[1] and power[2]<=self.__TPowerRev[2]
        backtime,backpower=self.__eval_back_time_power(plan,day,node)
        backtimeOK=backtime[0]<=self.__TTimeRev[0] and backtime[1]<=self.__TTimeRev[1] and backtime[2]<=self.__TTimeRev[2]
        backpowerOK=backpower[0]<=self.__TPowerRev[0] and backpower[1]<=self.__TPowerRev[1] and backpower[2]<=self.__TPowerRev[2]
        return timeOK and powerOK and backtimeOK and backpowerOK
    def __eval_plan_time_power(self,plan):
        time_usage,power_usage=[0.,0.,0.],[0.,0.,0.]
        i=0
        curnode=None
        for dayplan in plan:
            for node in dayplan:
                if curnode==None:
                    curnode=node
                else:
                    if curnode[0]==node[0]:
                        time=self.__task.loc[node[1],'time']
                        power=self.__task.loc[node[1],'power']
                    else:
                        time=self.__tmatrix.loc[self.__pointdfbackup.loc[curnode[0],'index'],self.__pointdfbackup.loc[node[0],'index']]+self.__task.loc[node[1],'time']
                        power=self.__pmatrix.loc[self.__pointdfbackup.loc[curnode[0],'index'],self.__pointdfbackup.loc[node[0],'index']]+self.__task.loc[node[1],'power']
                    time_usage[i]+=time
                    power_usage[i]+=power
                    curnode=node
            i+=1
        return time_usage,power_usage
    def __eval_back_time_power(self,plan,day,newnode):
        time_usage,power_usage=[0.,0.,0.],[0.,0.,0.]
        curnode=None
        end=False
        for node in plan[day]:
            if curnode==None:
                curnode=node
            elif curnode==newnode:
                time=self.__tmatrix.loc[self.__pointdfbackup.loc[curnode[0],'index'],self.__pointdfbackup.loc[self.__opoint[day+1][0],'index']]
                power=self.__pmatrix.loc[self.__pointdfbackup.loc[curnode[0],'index'],self.__pointdfbackup.loc[self.__opoint[day+1][0],'index']]
                end=True
            else:
                if curnode[0]==node[0]:
                    time=self.__task.loc[node[1],'time']
                    power=self.__task.loc[node[1],'power']
                else:
                    time=self.__tmatrix.loc[self.__pointdfbackup.loc[curnode[0],'index'],self.__pointdfbackup.loc[node[0],'index']]+self.__task.loc[node[1],'time']
                    power=self.__pmatrix.loc[self.__pointdfbackup.loc[curnode[0],'index'],self.__pointdfbackup.loc[node[0],'index']]+self.__task.loc[node[1],'power']
                time_usage[day]+=time
                power_usage[day]+=power
                curnode=node
                if end:
                    break
        return time_usage,power_usage
    def __cal_score(self,plan,newnode,newplan,w=(1.,0.)):
        time,power=self.__eval_plan_time_power(plan)
        newtime,newpower=self.__eval_plan_time_power(newplan)
        score=w[0]*(self.__task.loc[newnode[1],'revenue']/(sum(newtime)-sum(time)))+w[1]*(self.__task.loc[newnode[1],'revenue']/(sum(newpower)-sum(power)))
        return score
    def __find_best_pos(self,plan,node):
        bplan=None
        btaskno=None
        bscore=0.
        if node[1] in self.__daytaskindex[0]:
            days=(0,)
        elif node[1] in self.__daytaskindex[1]:
            days=(1,)
        elif node[1] in self.__daytaskindex[2]:
            days=(2,)
        elif node[1] in self.__daytaskindex[3]:
            days=(0,1)
        elif node[1] in self.__daytaskindex[4]:
            days=(1,2)
        elif node[1] in self.__daytaskindex[5]:
            days=(0,2)
        else:
            days=(0,1,2)
        for day in days:
            for i in range(1,len(plan[day])):
                newplan=deepcopy(plan)
                newplan[day].insert(i,node)
                if self.__check_plan(newplan,day,node):
                    score=self.__cal_score(plan,node,newplan)
                else:
                    score=0.
                if score>bscore:
                    bscore=score
                    bplan=deepcopy(newplan)
                    btaskno=node[1]
        return bplan,bscore,btaskno
    def __iter_task(self,plan,tasklist):
        plancopy=deepcopy(plan)
        bbplan=None
        bbtaskno=None
        bbscore=0.
        nodes=[]
        for taskno in tasklist:
            pts=self.__task.loc[taskno,'location']
            for pt in pts:
                ptno=self.__pointdf.loc[pt,'No']
                nodes.append((ptno,taskno))
        for node in nodes:
            bplan,bscore,btaskno=self.__find_best_pos(plancopy,node)
            if bscore>bbscore:
                bbscore=bscore
                bbplan=deepcopy(bplan)
                bbtaskno=btaskno
        return bbplan,bbtaskno
    def _proc_cont(self):
        if self.__Status==CONST.INPROCESS:
            contask=deepcopy(self.__contaskindex)
            i=1
            while i<=self.__Mincontinuous:
                plan,taskno=self.__iter_task(self.__res,contask)
                if plan==None:
                    raise MIPError('Cannot assign all continuous tasks')
                else:
                    self.__res=plan
                    contask.remove(taskno)
                    i+=1
        return None
    def _proc_req(self):
        if self.__Status==CONST.INPROCESS:
            reqtask=deepcopy(self.__reqtaskindex)
            while len(reqtask)>0:
                plan,taskno=self.__iter_task(self.__res,reqtask)
                if plan==None:
                    raise MIPError('Cannot assign all required tasks')
                else:
                    self.__res=plan
                    reqtask.remove(taskno)
        return None
    def _add_tag(self,days=(0,1)):
        if self.__Status==CONST.INPROCESS:
            if days[0]==0:
                if self.__current_state[0]==None:
                    self.__res[0].insert(1,(self.__pointdf.loc['探测起点1','No'],self.__task[self.__task['tag']=='12s'].index[0]))
                    self.__res[0].insert(-1,(self.__pointdf.loc['探测起点2','No'],self.__task[self.__task['tag']=='12e'].index[0]))
                if self.__current_state[0]=='12s':
                    self.__res[0].insert(-1,(self.__pointdf.loc['探测起点2','No'],self.__task[self.__task['tag']=='12e'].index[0]))
                elif self.__current_state[0] in ('12e','23s','23e'):
                    pass
            elif days[0]==1:
                if self.__current_state[0]==None:
                    self.__res[1].insert(1,(self.__pointdf.loc['探测起点1','No'],self.__task[self.__task['tag']=='12s'].index[0]))
                    self.__res[1].insert(-1,(self.__pointdf.loc['探测起点2','No'],self.__task[self.__task['tag']=='12e'].index[0]))
                if self.__current_state[0]=='12s':
                    self.__res[1].insert(-1,(self.__pointdf.loc['探测起点2','No'],self.__task[self.__task['tag']=='12e'].index[0]))
                elif self.__current_state[0] in ('12e','23s','23e'):
                    pass
            if days[1]==1:
                if self.__current_state[0] in (None,'12s','12e'):
                    self.__res[1].insert(1,(self.__pointdf.loc['探测起点2','No'],self.__task[self.__task['tag']=='23s'].index[0]))
                    self.__res[1].insert(-1,(self.__pointdf.loc['探测起点3','No'],self.__task[self.__task['tag']=='23e'].index[0]))
                elif self.__current_state[0]=='23s':
                    self.__res[1].insert(-1,(self.__pointdf.loc['探测起点3','No'],self.__task[self.__task['tag']=='23e'].index[0]))
                elif self.__current_state[0]=='23e':
                    pass
            elif days[1]==2:
                if self.__current_state[0] in (None,'12s','12e'):
                    self.__res[2].insert(1,(self.__pointdf.loc['探测起点2','No'],self.__task[self.__task['tag']=='23s'].index[0]))
                    self.__res[2].insert(-1,(self.__pointdf.loc['探测起点3','No'],self.__task[self.__task['tag']=='23e'].index[0]))
                elif self.__current_state[0]=='23s':
                    self.__res[2].insert(-1,(self.__pointdf.loc['探测起点3','No'],self.__task[self.__task['tag']=='23e'].index[0]))
                elif self.__current_state[0]=='23e':
                    pass
        return None
    def _proc_opt(self):
        if self.__Status==CONST.INPROCESS:
            opttask=deepcopy(self.__opttaskindex)
            status=True
            while len(self.__opttaskindex)>0 and status:
                plan,taskno=self.__iter_task(self.__res,opttask)
                if plan==None:
                    status=False
                else:
                    self.__res=plan
                    opttask.remove(taskno)
        return None
    def __cal_vals(self):
        time_usage,power_usage=self.__eval_plan_time_power(self.__res)
        time_efficiency=(sum(time_usage)+sum(self.__PKGTime)+sum(self.__CompTime))/(sum(self.__TTime)+sum(self.__PKGTime)+sum(self.__CompTime))*100
        power_efficiency=(sum(power_usage)+sum(self.__PKGPower)+sum(self.__CompPower))/(sum(self.__TPower)+sum(self.__PKGPower)+sum(self.__CompPower))*100
        res=self.__res[0][:-1]+self.__res[1][:-1]+self.__res[2][:-1]
        revenues=[self.__task.loc[no,'revenue'] for _,no in res]
        objVal=sum(revenues)
        return objVal,time_efficiency,power_efficiency
    def __format_number(self,number):
        number=float(number)
        format_str='{:.'+str(self.__decimal)+'f}'
        formatted_number=format_str.format(number)
        while formatted_number.endswith('0'):
            formatted_number=formatted_number[:-1]
        if formatted_number.endswith('.'):
            formatted_number=formatted_number[:-1]
        return formatted_number
    def _print_status(self,n=None):
        objVal,timeEff,powerEff=self.__cal_vals()
        objVal,timeEff,powerEff=[self.__format_number(number) for number in [objVal,timeEff,powerEff]]
        self.__Status=CONST.SUCCESS
        if n==None:
            print('Feasible objective value '+objVal+' found with '+timeEff+'% time-efficiency and '+powerEff+'% power-efficiency')
        else:
            print('Feasible objective value '+objVal+f' found for Occasion {n} with '+timeEff+'% time-efficiency and '+powerEff+'% power-efficiency')
        return None
    def iter_tags(self):
        if self.__completed_tag[0]==None:
            occasion=((0,1),(0,2),(1,2))
        elif self.__completed_tag[0]==0:
            if self.__completed_tag[1]==None:
                occasion=((0,1),(0,2))
            elif self.__completed_tag[1]==0:
                if self.__completed_tag[2]==None:
                    occasion=((0,1),(0,2))
                elif self.__completed_tag[2]==1:
                    occasion=((0,1),)
                elif self.__completed_tag[2]==2:
                    occasion=((0,2),)
        elif self.__completed_tag[0]==1:
            occasion=((1,2),)
        n=1
        btimeEff=0.
        bobjVal=0.
        bres=None
        errors=[False for i in range(len(occasion)*2)]
        errormessages=['' for i in range(len(occasion)*2)]
        
        for days in occasion:
            try:
                self._gen_init_plan(n*2-1)
                self._proc_tag(days)
                self._proc_cont()
                self._proc_req()
                self._proc_opt()
                self._add_tag(days)
                self._print_status(n)
                objVal1,timeEff1,_=self.__cal_vals()
                resbackup=deepcopy(self.__res)
            except MIPError as e:
                objVal1,timeEff1=0.,0.
                errors[(n-1)*2]=True
                errormessages[(n-1)*2]=e.message.lower()
                print(f'No feasible solution found for Occasion {n*2-1} because {e.message.lower()}')
            try:
                self._gen_init_plan(n*2)
                self._proc_tag(days)
                self._proc_req()
                self._proc_cont()
                self._proc_opt()
                self._add_tag(days)
                self._print_status(n)
                objVal2,timeEff2,_=self.__cal_vals()
            except MIPError as e:
                objVal2,timeEff2=0.,0.
                errors[(n-1)*2+1]=True
                errormessages[(n-1)*2+1]=e.message.lower()
                print(f'No feasible solution found for Occasion {n*2} because {e.message.lower()}')
            if objVal1>objVal2:
                self.__res=resbackup
                objVal=objVal1
                timeEff=timeEff1
            elif objVal1==objVal2 and timeEff1>timeEff2:
                self.__res=resbackup
                objVal=objVal1
                timeEff=timeEff1
            else:
                objVal=objVal2
                timeEff=timeEff2
            if objVal>bobjVal:
                bres=deepcopy(self.__res)
                btimeEff=timeEff
                bobjVal=objVal
            elif objVal==bobjVal and timeEff>btimeEff:
                bres=deepcopy(self.__res)
                btimeEff=timeEff
                bobjVal=objVal
            n+=1
        if errors.count(True)>=len(occasion)*2:
            errormessages=list(set(errormessages))
            if len(errormessages)==1:
                raise MIPError(f'No feasible solution found because {errormessages[0]}')
            elif len(errormessages)==2:
                raise MIPError(f'No feasible solution found because {errormessages[0]} and {errormessages[0]}')
            else:
                raise MIPError('No feasible solution found')
        print(f'Comparing objective values among {(n-1)*2-1} occasions')
        self.__res=bres
        self._print_status()
        return None
    def cal_route(self):
        pointdf=self.__pointdf.copy()
        pointdf.reset_index(inplace=True)
        pointdf.set_index('No',inplace=True)
        taskdf=self.__task.copy()
        taskdf.reset_index(inplace=True)
        taskdf.set_index('No',inplace=True)
        plan=self.__res.copy()
        n=1
        self.__plandf=[]
        for pli in range(3):
            no,action,location,coordinate,time,power,player1,player2,completed=[],[],[],[],[],[],[],[],[]
            if self.__res[pli][0][1]==self.__opoint[pli][1]:
                curpt=self.__res[pli][0][0]
            if list(self.__opoint[0]) in plan[0]:
                plan[0].remove(list(self.__opoint[0]))
            for i,k in plan[pli]:
                if i==curpt or ((i in (self.__opoint[i][0] for i in range(4))) and (curpt in (self.__opoint[i][0] for i in range(4)))):
                    no.append(n)
                    pti=pointdf.loc[i,'index']
                    ptcur=pointdf.loc[curpt,'index']
                    action.append(taskdf.loc[k,'name'])
                    location.append('探测起点' if pointdf.loc[i,'index'] in ['探测起点1','探测起点2','探测起点3','探测起点4'] else pointdf.loc[i,'index'])
                    coordinate.append(f"({self.__format_number(pointdf.loc[i,'X'])},{self.__format_number(pointdf.loc[i,'Y'])})")
                    time.append(float(self.__format_number(taskdf.loc[k,'time'])))
                    power.append(float(self.__format_number(taskdf.loc[k,'power'])))
                    player1.append('√')
                    player2.append('√')
                    completed.append(np.nan)
                    curpt=i
                    n+=1
                elif i!=curpt:
                    pti=pointdf.loc[i,'index']
                    ptcur=pointdf.loc[curpt,'index']
                    if pti!=ptcur:
                        no.append(n)
                        action.append(f"Travel from ({self.__format_number(pointdf.loc[curpt,'X'])},{self.__format_number(pointdf.loc[curpt,'X'])}) to ({self.__format_number(pointdf.loc[i,'X'])},{self.__format_number(pointdf.loc[i,'Y'])})")
                        ttime=self.__tmatrix.loc[pointdf.loc[curpt,'index']][pointdf.loc[i,'index']]
                        time.append(float(self.__format_number(ttime)))
                        tpower=self.__pmatrix.loc[pointdf.loc[curpt,'index']][pointdf.loc[i,'index']]
                        power.append(float(self.__format_number(tpower)))
                        location.append(f"{'探测起点' if pointdf.loc[curpt,'index'] in ['探测起点1','探测起点2','探测起点3','探测起点4'] else pointdf.loc[curpt,'index']}→{'探测起点' if pointdf.loc[i,'index'] in ['探测起点1','探测起点2','探测起点3','探测起点4'] else pointdf.loc[i,'index']}")
                        coordinate.append(f"({self.__format_number(pointdf.loc[curpt,'X'])},{self.__format_number(pointdf.loc[curpt,'Y'])})→({self.__format_number(pointdf.loc[i,'X'])},{self.__format_number(pointdf.loc[i,'Y'])})")
                        player1.append('√')
                        player2.append('√')
                        completed.append(np.nan)
                        n+=1
                    no.append(n)
                    action.append(taskdf.loc[k,'name'])
                    location.append('探测起点' if pointdf.loc[i,'index'] in ['探测起点1','探测起点2','探测起点3','探测起点4'] else pointdf.loc[i,'index'])
                    coordinate.append(f"({self.__format_number(pointdf.loc[i,'X'])},{self.__format_number(pointdf.loc[i,'Y'])})")
                    time.append(float(self.__format_number(taskdf.loc[k,'time'])))
                    power.append(float(self.__format_number(taskdf.loc[k,'power'])))
                    player1.append('√')
                    player2.append('√')
                    completed.append(np.nan)
                    curpt=i
                    n+=1
            no.pop()
            action.pop()
            location.pop()
            coordinate.pop()
            time.pop()
            power.pop()
            player1.pop()
            player2.pop()
            completed.pop()
            self.__plandf.append(pd.DataFrame({'No':no,'action':action,'location':location,'coordinate':coordinate,'time':time,'power':power,'1':player1,'2':player2,'completed':completed}))
        for j in range(3):
            if (self.__res[j][0][1] in [self.__opoint[i][1] for i in range(4)]) or (self.__res[j][0][0] in [self.__opoint[i][0] for i in range(4)]):
                self.__plandf[j]=self.__plandf[j][1:]
        return None
    def __gen_packagedf(self,package,tag):
        pak=self.__package[package['tag']==tag]
        no=[i for i in range(pak.shape[0])]
        location=f"({self.__format_number(self.__pointdf.loc['探测起点1','X'])},{self.__format_number(self.__pointdf.loc['探测起点1','X'])})"
        coord=[location for i in range(pak.shape[0])]
        loc=['探测起点' for i in range(pak.shape[0])]
        player=['√' for i in range(pak.shape[0])]
        completed=[np.nan for i in range(pak.shape[0])]
        df=pd.DataFrame({'No':no,'action':pak.loc[:,'name'].values,'location':loc,'coordinate':coord,'time':pak.loc[:,'time'].values,'power':pak.loc[:,'power'].values,'1':player,'2':player,'completed':completed})
        return df
    def add_package(self):
        if self.__completed_schedule[0].shape[0]>0:
            self.__plandf[0]=pd.concat([self.__completed_schedule[0],self.__plandf[0]])
        if self.__completed_schedule[1].shape[0]>0:
            self.__plandf[1]=pd.concat([self.__completed_schedule[1],self.__plandf[1]])
        if self.__completed_schedule[2].shape[0]>0:
            self.__plandf[2]=pd.concat([self.__completed_schedule[2],self.__plandf[2]])
        plan=self.__plandf.copy()
        isindf12=[False,False,False]
        isindf23=[False,False,False]
        df=[False,False,False]
        df0=pd.DataFrame({'No':[np.nan],'action':['Begin of Day1'],'location':[np.nan],'coordinate':[np.nan],'time':[np.nan],'power':[np.nan],'1':[np.nan],'2':[np.nan],'completed':[np.nan]})
        for i in range(3):
            isindf12[i]=self.__plandf[i].isin(self.__task[self.__task['tag']=='12s']['name'].values).any().any()
            isindf23[i]=self.__plandf[i].isin(self.__task[self.__task['tag']=='23s']['name'].values).any().any()
        pakdf=[[[self.__gen_packagedf(self.__package,'D1ss'),self.__gen_packagedf(self.__package,'D1se')],[self.__gen_packagedf(self.__package,'D1es'),self.__gen_packagedf(self.__package,'D1ee')]],
                [[self.__gen_packagedf(self.__package,'D2ss'),self.__gen_packagedf(self.__package,'D2se')],[self.__gen_packagedf(self.__package,'D2es'),self.__gen_packagedf(self.__package,'D2ee')]],
                [[self.__gen_packagedf(self.__package,'D3ss'),self.__gen_packagedf(self.__package,'D3se')],[self.__gen_packagedf(self.__package,'D3es'),self.__gen_packagedf(self.__package,'D3ee')]]]
        for i in range(3):
            if isindf12[i]:
                if self.__plandf[i].shape[0]==2:
                    temp=pd.concat([self.__plandf[i].iloc[:1,:],pakdf[i][0][1],pakdf[i][1][0],self.__plandf[i].iloc[-1:,:]])
                elif self.__plandf[i].shape[0]>2:
                    temp=pd.concat([self.__plandf[i].iloc[:1,:],pakdf[i][0][1],self.__plandf[i].iloc[1:-1,:],pakdf[i][1][0],self.__plandf[i].iloc[-1:,:]])
                time=sum(temp['time'])
                if time<=self.__12gap:
                    adddf=pd.DataFrame({'No':[np.nan],'action':[f'Wait for {self.__format_number(self.__12gap-time)}s'],'location':['探测起点'],'coordinate':[f"({self.__format_number(self.__pointdf.loc['探测起点1','X'])},{self.__format_number(self.__pointdf.loc['探测起点1','X'])})"],'time':[float(self.__format_number(self.__12gap-time))],'power':[0],'1':['√'],'2':['√']})
                    if self.__plandf[i].shape[0]==2:
                        plan[i]=pd.concat([pakdf[i][0][0],self.__plandf[i].iloc[:1,:],pakdf[i][0][1],pakdf[i][1][0],adddf,self.__plandf[i].iloc[-1:,:],pakdf[i][1][1]])
                    elif self.__plandf[i].shape[0]>2:
                        plan[i]=pd.concat([pakdf[i][0][0],self.__plandf[i].iloc[:1,:],pakdf[i][0][1],self.__plandf[i].iloc[1:-1,:],pakdf[i][1][0],adddf,self.__plandf[i].iloc[-1:,:],pakdf[i][1][1]])
                else:
                    if self.__plandf[i].shape[0]==2:
                        plan[i]=pd.concat([pakdf[i][0][0],self.__plandf[i].iloc[:1,:],pakdf[i][0][1],pakdf[i][1][0],self.__plandf[i].iloc[-1:,:],pakdf[i][1][1]])
                    elif self.__plandf[i].shape[0]>2:
                        plan[i]=pd.concat([pakdf[i][0][0],self.__plandf[i].iloc[:1,:],pakdf[i][0][1],self.__plandf[i].iloc[1:-1,:],pakdf[i][1][0],self.__plandf[i].iloc[-1:,:],pakdf[i][1][1]])
                time='sum: {}'.format(self.__format_number(sum(plan[i]['time'])))
                power='sum: {}'.format(self.__format_number(sum(plan[i]['power'])))
                df[i]=pd.DataFrame({'No':[np.nan],'action':['xxx'],'location':[np.nan],'coordinate':[np.nan],'time':[time],'power':[power],'1':[np.nan],'2':[np.nan],'completed':[np.nan]})
            elif isindf23[i]:
                if self.__plandf[i].shape[0]==2:
                    temp=pd.concat([self.__plandf[i].iloc[:1,:],pakdf[i][0][1],pakdf[i][1][0],self.__plandf[i].iloc[-1:,:]])
                elif self.__plandf[i].shape[0]>2:
                    temp=pd.concat([self.__plandf[i].iloc[:1,:],pakdf[i][0][1],self.__plandf[i].iloc[1:-1,:],pakdf[i][1][0],self.__plandf[i].iloc[-1:,:]])
                time=sum(temp['time'])
                if time<=self.__23gap:
                    adddf=pd.DataFrame({'No':[np.nan],'action':[f'Wait for {self.__format_number(self.__23gap-time)}s'],'location':['探测起点'],'coordinate':[f"({self.__format_number(self.__pointdf.loc['探测起点1','X'])},{self.__format_number(self.__pointdf.loc['探测起点1','X'])})"],'time':[float(self.__format_number(self.__23gap-time))],'power':[0],'1':['√'],'2':['√']})
                    if self.__plandf[i].shape[0]==2:
                        plan[i]=pd.concat([pakdf[i][0][0],self.__plandf[i].iloc[:1,:],pakdf[i][0][1],pakdf[i][1][0],adddf,self.__plandf[i].iloc[-1:,:],pakdf[i][1][1]])
                    elif self.__plandf[i].shape[0]>2:
                        plan[i]=pd.concat([pakdf[i][0][0],self.__plandf[i].iloc[:1,:],pakdf[i][0][1],self.__plandf[i].iloc[1:-1,:],pakdf[i][1][0],adddf,self.__plandf[i].iloc[-1:,:],pakdf[i][1][1]])
                else:
                    if self.__plandf[i].shape[0]==2:
                        plan[i]=pd.concat([pakdf[i][0][0],self.__plandf[i].iloc[:1,:],pakdf[i][0][1],pakdf[i][1][0],self.__plandf[i].iloc[-1:,:],pakdf[i][1][1]])
                    elif self.__plandf[i].shape[0]>2:
                        plan[i]=pd.concat([pakdf[i][0][0],self.__plandf[i].iloc[:1,:],pakdf[i][0][1],self.__plandf[i].iloc[1:-1,:],pakdf[i][1][0],self.__plandf[i].iloc[-1:,:],pakdf[i][1][1]])
                time='sum: {}'.format(self.__format_number(sum(plan[i]['time'])))
                power='sum: {}'.format(self.__format_number(sum(plan[i]['power'])))
                df[i]=pd.DataFrame({'No':[np.nan],'action':['xxx'],'location':[np.nan],'coordinate':[np.nan],'time':[time],'power':[power],'1':[np.nan],'2':[np.nan],'completed':[np.nan]})
            else:
                pakdf[i][0]=pd.concat(pakdf[i][0])
                pakdf[i][1]=pd.concat(pakdf[i][1])
                plan[i]=pd.concat([pakdf[i][0],self.__plandf[i],pakdf[i][1]])
                time='sum: {}'.format(self.__format_number(sum(plan[i]['time'])))
                power='sum: {}'.format(self.__format_number(sum(plan[i]['power'])))
                df[i]=pd.DataFrame({'No':[np.nan],'action':['xxx'],'location':[np.nan],'coordinate':[np.nan],'time':[time],'power':[power],'1':[np.nan],'2':[np.nan],'completed':[np.nan]})
        df[0]['action']='Break between Day1 and Day2'
        df[1]['action']='Break between Day2 and Day3'
        df[2]['action']='End of Day3'
        self.__plandf=pd.concat([df0,plan[0],df[0],plan[1],df[1],plan[2],df[2]])
        self.__plandf.reset_index(drop=True,inplace=True)
        index0=self.__plandf[self.__plandf['action']=='Begin of Day1'].index[0]+2
        index1=self.__plandf[self.__plandf['action']=='Break between Day1 and Day2'].index[0]+2
        index2=self.__plandf[self.__plandf['action']=='Break between Day2 and Day3'].index[0]+2
        index3=self.__plandf[self.__plandf['action']=='End of Day3'].index[0]+2
        self.__voidindex=[int(index0),int(index1),int(index2),int(index3)]
        no=[np.nan]+list(range(1,index1-2))+[np.nan]+list(range(index1-2,index2-3))+[np.nan]+list(range(index2-3,index3-4))+[np.nan]
        self.__plandf['No']=no
        return None
    def write_excel(self):
        with pd.ExcelWriter(self.__outputpath) as writer:
            self.__plandf.to_excel(writer,index=False)
            worksheet=writer.sheets[list(writer.sheets.keys())[0]]
            for column_cells in worksheet.columns:
                lengthlist=[]
                for cell in column_cells:
                    length=0
                    for char in str(cell.value):
                        if char>=u'\u4e00' and char<=u'\u9fff':
                            length+=2
                        else:
                            length+=1
                    lengthlist.append(length)
                length=max(max(lengthlist),5)
                worksheet.column_dimensions[column_cells[0].column_letter].width=length+2
            for row in self.__voidindex:
                for cell in worksheet[row]:
                    cell.fill=PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
        timecost=self.__format_number(time.time()-self.__starttime)
        print('Process finished in '+timecost+' seconds')
        return None
    def schedule_frame(self):
        return self.__plandf.copy()
    def run(self):
        self.test_IO()
        self.read_info()
        self.read_task()
        self.read_package()
        self.read_point()
        self.gen_void_point()
        self.check_remote()
        self.divide_task()
        self.drop_O()
        self.update_matrix()
        self.read_schedule()
        self.update_task()
        self.update_package()
        self.iter_tags()
        self.cal_route()
        self.add_package()
        if self.__write_output:
            self.write_excel()
        return None
class heuristic_back(object):
    def __init__(self,workPath=None,decimal=5,timePathRevised=None,powerPathRevised=None,
                 infoPath='info.xlsx',taskPath='task.xlsx',packPath='package.xlsx',
                 pointPath='point.xlsx',distancePath='distance.xlsx',timePath='time.xlsx',
                 powerPath='power.xlsx',schedulePath='schedule.xlsx',outputPath='schedule-revised.xlsx',
                 dataFrames=None,scheduleFrame=None,writeOutput=True):
        print('Process initiated')
        self.__starttime=time.time()
        self.__decimal=decimal
        if workPath==None:
            self.__infoPath,self.__taskPath,self.__pakpath=infoPath,taskPath,packPath
            self.__pointpath,self.__distancepath,self.__timepath,self.__powerpath=pointPath,distancePath,timePath,powerPath
        else:
            self.__infoPath,self.__taskPath,self.__pakpath=os.path.join(workPath,infoPath),os.path.join(workPath,taskPath),os.path.join(workPath,packPath)
            self.__pointpath,self.__distancepath,self.__timepath,self.__powerpath=os.path.join(workPath,pointPath),os.path.join(workPath,distancePath),os.path.join(workPath,timePath),os.path.join(workPath,powerPath)
        self.__timepathrevised,self.__powerpathrevised=timePathRevised,powerPathRevised
        self.__schedulepath=schedulePath
        self.__outputpath=outputPath
        self.__dataframes=dataFrames or {}
        self.__scheduleframe=scheduleFrame
        self.__write_output=writeOutput
        self.__Status=CONST.CREATED
    def __trans_time(self,time):
        newtime=time*2
        return newtime
    def __trans_power(self,power):
        newpower=power*1.5
        return newpower
    def test_IO(self):
        if self.__dataframes:
            missing=[key for key in ['info','task','package','point','distance','time','power'] if key not in self.__dataframes]
            if missing:
                raise ValueError(f'Missing DataFrame inputs: {missing}')
            if self.__scheduleframe is None:
                raise ValueError('Missing scheduleFrame input')
            return None
        FileNotFound=[]
        if not os.path.exists(self.__infoPath):
            FileNotFound.append(self.__infoPath)
        if not os.path.exists(self.__taskPath):
            FileNotFound.append(self.__taskPath)
        if not os.path.exists(self.__pakpath):
            FileNotFound.append(self.__pakpath)
        if not os.path.exists(self.__pointpath):
            FileNotFound.append(self.__pointpath)
        if not os.path.exists(self.__distancepath):
            FileNotFound.append(self.__distancepath)
        if not os.path.exists(self.__timepath):
            FileNotFound.append(self.__timepath)
        if not os.path.exists(self.__powerpath):
            FileNotFound.append(self.__powerpath)
        if not os.path.exists(self.__schedulepath):
            FileNotFound.append(self.__schedulepath)
        if self.__timepathrevised != None:
            if not os.path.exists(self.__timepathrevised):
                FileNotFound.append(self.__timepathrevised)
        if self.__powerpathrevised != None:
            if not os.path.exists(self.__powerpathrevised):
                FileNotFound.append(self.__powerpathrevised)
        if FileNotFound!=[]:
            raise FileNotFoundError('No such file: {}'.format(' '.join(FileNotFound)))
        notPermitted=False
        if self.__write_output and os.path.exists(self.__outputpath):
            try:
                pd.read_excel(self.__outputpath)
            except PermissionError:
                notPermitted=True
        if notPermitted:
            raise PermissionError(f'Permission denied: {self.__outputpath}')
        return None
    def read_info(self):
        info=self.__dataframes.get('info')
        info=info.copy() if info is not None else pd.read_excel(self.__infoPath)
        self.__MDistance=info['max-distance'][0]
        self.__TTime=list(map(float,info['total-time/day'][0].split(';')))
        self.__TPower=list(map(float,info['total-power/day'][0].split(';')))
        self.__TTimebackup=deepcopy(self.__TTime)
        self.__TPowerbackup=deepcopy(self.__TPower)
        self.__Mincontinuous=info['min-continuous'][0]
        self.__12gap=info['12-gap'][0]
        self.__23gap=info['23-gap'][0]
        return None
    def __read_matrix(self,path,key=None):
        matrix=self.__dataframes.get(key) if key is not None else None
        if matrix is not None:
            matrix=matrix.copy()
        else:
            matrix=pd.read_excel(path)
            matrix.set_index(matrix.columns[0],inplace=True)
            matrix.index.rename(None,inplace=True)
        return matrix
    def read_task(self):
        task=self.__dataframes.get('task')
        self.__task=task.copy() if task is not None else pd.read_excel(self.__taskPath)
        self.__dmatrix=self.__read_matrix(self.__distancepath,'distance')
        self.__dmatrix.replace(np.inf,self.__MDistance*3,inplace=True)
        np.fill_diagonal(self.__dmatrix.values,0)
        self.__tmatrix=self.__read_matrix(self.__timepath,'time')
        self.__tmatrix.replace(np.inf,max(self.__TTime),inplace=True)
        np.fill_diagonal(self.__tmatrix.values,0)
        self.__pmatrix=self.__read_matrix(self.__powerpath,'power')
        self.__pmatrix.replace(np.inf,max(self.__TPower),inplace=True)
        np.fill_diagonal(self.__pmatrix.values,0)
        return None
    def read_package(self):
        package=self.__dataframes.get('package')
        self.__package=package.copy() if package is not None else pd.read_excel(self.__pakpath)
        time=[sum(self.__package[self.__package['tag']==tag]['time']) for tag in ['D1ss','D1se','D1es','D1ee','D2ss','D2se','D2es','D2ee','D3ss','D3se','D3es','D3ee']]
        power=[sum(self.__package[self.__package['tag']==tag]['power']) for tag in ['D1ss','D1se','D1es','D1ee','D2ss','D2se','D2es','D2ee','D3ss','D3se','D3es','D3ee']]
        self.__PKGTime=[time[0]+time[1]+time[2]+time[3],time[4]+time[5]+time[6]+time[7],time[8]+time[9]+time[10]+time[11]]
        self.__PKGPower=[power[0]+power[1]+power[2]+power[3],power[4]+power[5]+power[6]+power[7],power[8]+power[9]+power[10]+power[11]]
        self.__TTime=[self.__TTime[i]-self.__PKGTime[i] for i in range(3)]
        self.__TPower=[self.__TPower[i]-self.__PKGPower[i] for i in range(3)]
        return None
    def read_point(self):
        point=self.__dataframes.get('point')
        self.__pointdf=point.copy() if point is not None else pd.read_excel(self.__pointpath)
        self.__pointdf.set_index(self.__pointdf.columns[0],inplace=True)
        self.__pointdf.index.rename(None,inplace=True)
        return None
    def __add_void_matrix(self,matrix):
        voidrow=[matrix.loc['探测起点',:].values.tolist() for i in range(4)]
        voidrowdf=pd.DataFrame(voidrow,index=['探测起点1','探测起点2','探测起点3','探测起点4'],columns=matrix.columns)
        matrix=pd.concat([voidrowdf,matrix])
        matrix.drop(['探测起点'],inplace=True)
        voidcol=np.array([matrix.loc[:,'探测起点'].values.tolist() for i in range(4)]).T
        voidcoldf=pd.DataFrame(voidcol,columns=['探测起点1','探测起点2','探测起点3','探测起点4'],index=matrix.index)
        matrix=pd.concat([voidcoldf,matrix],axis=1)
        matrix.drop(['探测起点'],axis=1,inplace=True)
        for i,j in product(['探测起点1','探测起点2','探测起点3','探测起点4'],['探测起点1','探测起点2','探测起点3','探测起点4']):
            if i!=j:
                matrix.loc[i,j]=0
        return matrix
    def gen_void_point(self):
        voidpoint=pd.DataFrame({'X':[self.__pointdf.loc['探测起点','X'],self.__pointdf.loc['探测起点','X'],self.__pointdf.loc['探测起点','X'],self.__pointdf.loc['探测起点','X']],
                                'Y':[self.__pointdf.loc['探测起点','Y'],self.__pointdf.loc['探测起点','Y'],self.__pointdf.loc['探测起点','Y'],self.__pointdf.loc['探测起点','Y']],
                                '备注':['虚拟原点：第一天出发','虚拟原点：第一天返回第二天出发','虚拟原点：第二天返回第三天出发','虚拟原点：第三天返回']},
                               index=['探测起点1','探测起点2','探测起点3','探测起点4'])
        self.__pointdf=pd.concat([voidpoint,self.__pointdf])
        self.__pointdf.drop(['探测起点'],inplace=True)
        self.__pointdf['No']=range(self.__pointdf.shape[0])
        self.__dmatrix=self.__add_void_matrix(self.__dmatrix)
        self.__tmatrix=self.__add_void_matrix(self.__tmatrix)
        self.__pmatrix=self.__add_void_matrix(self.__pmatrix)
        self.__task['location']=self.__task['location'].replace('探测起点','探测起点1,探测起点2,探测起点3,探测起点4')
        self.__task['location']=self.__task['location'].fillna(','.join(self.__pointdf.index.values))
        location=self.__task['location'].values
        pool=[]
        for i in location:
            if pd.isna(i):
                pool.append(np.nan)
            else:
                split=list(i.split(','))
                pool.append(split)
        self.__task['location']=pd.Series(pool)
        voidtaskdf=pd.DataFrame({'No':[self.__task.shape[0],self.__task.shape[0]+1,self.__task.shape[0]+2,self.__task.shape[0]+3],
                                 'name':['void1','void2','void3','void4'],
                                 'revenue':[0,0,0,0],
                                 'location':[['探测起点1'],['探测起点2'],['探测起点3'],['探测起点4']],
                                 'day':[np.nan,np.nan,np.nan,np.nan],
                                 'time':[0,0,0,0],
                                 'power':[0,0,0,0],
                                 'required':[False,False,False,False],
                                 'continuous':[False,False,False,False],
                                 'remote':[False,False,False,False],
                                 'exceptO':[False,False,False,False],
                                 'tag':[np.nan,np.nan,np.nan,np.nan]},
                                index=[self.__task.shape[0],self.__task.shape[0]+1,self.__task.shape[0]+2,self.__task.shape[0]+3])
        self.__task=pd.concat([self.__task,voidtaskdf])
        self.__opoint=((self.__pointdf.loc['探测起点1','No'],self.__task[self.__task['name']=='void1'].index[0]),
                       (self.__pointdf.loc['探测起点2','No'],self.__task[self.__task['name']=='void2'].index[0]),
                       (self.__pointdf.loc['探测起点3','No'],self.__task[self.__task['name']=='void3'].index[0]),
                       (self.__pointdf.loc['探测起点4','No'],self.__task[self.__task['name']=='void4'].index[0]))
        self.__pointdfbackup=self.__pointdf.copy()
        self.__pointdfbackup.reset_index(inplace=True)
        self.__pointdfbackup.set_index('No',inplace=True)
        return None
    def __cartesian_to_polar(self,x,y,O):
        r=np.sqrt((x-O[0])**2+(y-O[1])**2)
        theta=np.arctan2(y-O[1],x-O[0])
        return r,theta
    def __polar_to_cartesian(self,r,theta,O):
        x=r*np.cos(theta)+O[0]
        y=r*np.sin(theta)+O[1]
        return x,y
    def __cal_distance(self,point1,point2):
        dis=((point1[0]-point2[0])**2+(point1[1]-point2[1])**2)**0.5
        return dis
    def __from_distance(self,distance):
        pt=(distance,0)
        return pt
    def check_remote(self):
        remoteindex=self.__task[self.__task['remote']==True].index.to_list()
        points=[]
        for index in remoteindex:
            pts=[]
            for point in self.__task.loc[index,'location']:
                if self.__dmatrix.loc['探测起点1',point]>=self.__MDistance:
                    points.append(point)
                    pts.append(point)
            self.__task.at[index,'location']=pts
        if len(points)==0:
            start=0
            path='new-point.xlsx'
            sortedseries=self.__dmatrix.sort_values(by=['探测起点1'],ascending=False).loc['探测起点1',:]
            while sortedseries[start]>=self.__MDistance:
                start+=1
            pts=self.__dmatrix.sort_values(by=['探测起点1'],ascending=False).iloc[start:start+5,:].index.values
            r,_=self.__cartesian_to_polar(self.__from_distance(self.__MDistance)[0],self.__from_distance(self.__MDistance)[1],(0,0))
            pts=[self.__cartesian_to_polar(self.__pointdf.loc[pt,'X'],self.__pointdf.loc[pt,'Y'],(self.__pointdf.loc['探测起点1','X'],self.__pointdf.loc['探测起点1','Y'])) for pt in pts]
            pts=[self.__polar_to_cartesian(r,pt[1],(self.__pointdf.loc['探测起点1','X'],self.__pointdf.loc['探测起点1','Y'])) for pt in pts]
            newpt=pd.DataFrame(pts,columns=['X','Y'],index=pd.Index(['新最远探测点1','新最远探测点2','新最远探测点3','新最远探测点4','新最远探测点5'],name='name'))
            raise MaxDistanceError(f'No point meets the max distance requirement. Recommended points have been generated in {path}',newpt,path)
        return None
    def divide_task(self):
        self.__reqtaskindex=self.__task[self.__task['required']==True].index.to_list()
        self.__opttaskindex=self.__task[self.__task['required']==False].index.to_list()
        self.__daytaskindex=[self.__task[self.__task['day']==1].index.to_list(),
                             self.__task[self.__task['day']==2].index.to_list(),
                             self.__task[self.__task['day']==3].index.to_list(),
                             self.__task[self.__task['day']=='1,2'].index.to_list(),
                             self.__task[self.__task['day']=='2,3'].index.to_list(),
                             self.__task[self.__task['day']=='1,3'].index.to_list()]
        self.__remtaskindex=self.__task[self.__task['remote']==True].index.to_list()
        self.__tagtaskindex=[self.__task[self.__task['tag']=='12s'].index.to_list(),
                             self.__task[self.__task['tag']=='12e'].index.to_list(),
                             self.__task[self.__task['tag']=='23s'].index.to_list(),
                             self.__task[self.__task['tag']=='23e'].index.to_list()]
        self.__contaskindex=self.__task[self.__task['continuous']==True].index.to_list()
        tagtask=[]
        for i in range(len(self.__tagtaskindex)):
            tagtask.append(*self.__tagtaskindex[i])
        for index in self.__reqtaskindex.copy():
            if index in self.__contaskindex:
                self.__reqtaskindex.remove(index)
            if index in tagtask:
                self.__reqtaskindex.remove(index)
        for index in self.__opttaskindex.copy():
            if index in self.__contaskindex:
                self.__opttaskindex.remove(index)
            if index in tagtask:
                self.__opttaskindex.remove(index)
            if index in [self.__opoint[0][1],self.__opoint[1][1],self.__opoint[2][1],self.__opoint[3][1]]:
                self.__opttaskindex.remove(index)
        self.__noOtaskindex=self.__task[self.__task['exceptO']==True].index.to_list()
        return None
    def drop_O(self):
        for i in self.__noOtaskindex:
            self.__task.loc[i,'location'].remove('探测起点1')
            self.__task.loc[i,'location'].remove('探测起点2')
            self.__task.loc[i,'location'].remove('探测起点3')
            self.__task.loc[i,'location'].remove('探测起点4')
        return None
    def update_matrix(self):
        if self.__timepathrevised != None:
            self.__tmatrix=self.__read_matrix(self.__timepathrevised)
            self.__tmatrix.replace(np.inf,max(self.__TTime),inplace=True)
            np.fill_diagonal(self.__tmatrix.values,0)
            self.__tmatrix=self.__add_void_matrix(self.__tmatrix)
        else:
            self.__tmatrix=self.__tmatrix.map(self.__trans_time)
        if self.__powerpathrevised != None:
            self.__pmatrix=self.__read_matrix(self.__powerpathrevised)
            self.__pmatrix.replace(np.inf,max(self.__TPower),inplace=True)
            np.fill_diagonal(self.__pmatrix.values,0)
            self.__pmatrix=self.__add_void_matrix(self.__pmatrix)
        else:
            self.__pmatrix=self.__pmatrix.map(self.__trans_power)
        return None
    def __format_number(self,number):
        number=float(number)
        format_str='{:.'+str(self.__decimal)+'f}'
        formatted_number=format_str.format(number)
        while formatted_number.endswith('0'):
            formatted_number=formatted_number[:-1]
        if formatted_number.endswith('.'):
            formatted_number=formatted_number[:-1]
        return formatted_number
    def read_schedule(self):
        print('Solving feasible plan')
        taskno=self.__task.iloc[-1,0]+1
        points=['探测起点2','探测起点3','探测起点4']
        backtask=pd.DataFrame({'No':[taskno],'name':['紧急返回原点'],'revenue':[0],'location':[points],'day':[np.nan],'time':[0],'power':[0],'required':[False],'continuous':[False],'remote':[False],'exceptO':[False],'tag':[np.nan]},index=[taskno])
        self.__task=pd.concat([self.__task,backtask])
        schedule=self.__scheduleframe.copy() if self.__scheduleframe is not None else pd.read_excel(self.__schedulepath)
        index=schedule[schedule['No'].isna()].index.to_list()
        schedule=[schedule.iloc[index[0]+1:index[1],:],schedule.iloc[index[1]+1:index[2],:],schedule.iloc[index[2]+1:index[3],:]]
        self.__completed_schedule=[schedule[0][schedule[0]['completed'].notna()],schedule[1][schedule[1]['completed'].notna()],schedule[2][schedule[2]['completed'].notna()]]
        self.__res=[[],[],[]]
        index=1
        for day in range(3):
            if self.__completed_schedule[day].shape[0]==schedule[day].shape[0]:
                self.__res[day]=deepcopy(self.__completed_schedule[day])
            elif self.__completed_schedule[day].shape[0]==0:
                self.__res[day]=deepcopy(schedule[day])
            else:
                self.__res[day]=deepcopy(self.__completed_schedule[day])
                loc=self.__res[day].iloc[-1,:]['coordinate']
                if loc.count('→')==0:
                    X,Y=list(loc.split(','))
                elif loc.count('→')==1:
                    X,Y=list(loc.split(',')[1].split(','))
                X,Y=float(X[1:]),float(Y[:-1])
                lastptno=self.__find_point(X,Y)
                time_usage=sum(self.__res[day]['time'])
                power_usage=sum(self.__res[day]['power'])
                ttime=self.__tmatrix.loc[self.__pointdfbackup.loc[lastptno,'index'],points[day]]
                tpower=self.__pmatrix.loc[self.__pointdfbackup.loc[lastptno,'index'],points[day]]
                if ttime+time_usage<=self.__TTimebackup[day] and tpower+power_usage<=self.__TPowerbackup[day]:
                    no,action,location,coordinate,time,power,player1,player2,completed=[],[],[],[],[],[],[],[],[]
                    no.append(0)
                    action.append(f"Travel from ({self.__format_number(self.__pointdfbackup.loc[lastptno,'X'])},{self.__format_number(self.__pointdfbackup.loc[lastptno,'X'])}) to ({self.__format_number(self.__pointdf.loc[points[day],'X'])},{self.__format_number(self.__pointdf.loc[points[day],'Y'])})")
                    time.append(float(self.__format_number(ttime)))
                    power.append(float(self.__format_number(tpower)))
                    location.append(f"{'探测起点' if self.__pointdfbackup.loc[lastptno,'index'] in ['探测起点1','探测起点2','探测起点3','探测起点4'] else self.__pointdfbackup.loc[lastptno,'index']}→{'探测起点' if points[day] in ['探测起点1','探测起点2','探测起点3','探测起点4'] else points[day]}")
                    coordinate.append(f"({self.__format_number(self.__pointdfbackup.loc[lastptno,'X'])},{self.__format_number(self.__pointdfbackup.loc[lastptno,'X'])})→({self.__format_number(self.__pointdf.loc[points[day],'X'])},{self.__format_number(self.__pointdf.loc[points[day],'Y'])})")
                    player1.append('√')
                    player2.append('√')
                    completed.append(np.nan)
                    no.append(0)
                    action.append(self.__task.loc[taskno,'name'])
                    location.append('探测起点' if points[day] in ['探测起点1','探测起点2','探测起点3','探测起点4'] else points[day])
                    coordinate.append(f"({self.__format_number(self.__pointdf.loc[points[day],'X'])},{self.__format_number(self.__pointdf.loc[points[day],'Y'])})")
                    time.append(0)
                    power.append(0)
                    player1.append('√')
                    player2.append('√')
                    completed.append(np.nan)
                    self.__res[day]=pd.concat([self.__res[day],pd.DataFrame({'No':no,'action':action,'location':location,'time':time,'power':power,'1':player1,'2':player2,'completed':completed})])
                else:
                    raise MIPError('Not enough time or power to make it back to depot')
            self.__res[day]['No']=list(range(index,index+self.__res[day].shape[0]))
            index+=self.__res[day].shape[0]
        time=['sum: {}'.format(self.__format_number(sum(self.__res[i]['time']))) for i in range(3)]
        power=['sum: {}'.format(self.__format_number(sum(self.__res[i]['power']))) for i in range(3)]
        df0=pd.DataFrame({'No':[np.nan],'action':['Begin of Day1'],'location':[np.nan],'coordinate':[np.nan],'time':[np.nan],'power':[np.nan],'1':[np.nan],'2':[np.nan],'completed':[np.nan]})
        df1=pd.DataFrame({'No':[np.nan],'action':['Break between Day1 and Day2'],'location':[np.nan],'coordinate':[np.nan],'time':[time[0]],'power':[power[0]],'1':[np.nan],'2':[np.nan],'completed':[np.nan]})
        df2=pd.DataFrame({'No':[np.nan],'action':['Break between Day2 and Day3'],'location':[np.nan],'coordinate':[np.nan],'time':[time[1]],'power':[power[1]],'1':[np.nan],'2':[np.nan],'completed':[np.nan]})
        df3=pd.DataFrame({'No':[np.nan],'action':['End of Day3'],'location':[np.nan],'coordinate':[np.nan],'time':[time[2]],'power':[power[2]],'1':[np.nan],'2':[np.nan],'completed':[np.nan]})
        self.__plandf=pd.concat([df0,self.__res[0],df1,self.__res[1],df2,self.__res[2],df3])
        self.__plandf.reset_index(drop=True,inplace=True)
        self.__plandf.index.rename(None,inplace=True)
        self.__voidindex=[self.__plandf[self.__plandf['action']=='Begin of Day1'].index.to_list()[0]+2,
                          self.__plandf[self.__plandf['action']=='Break between Day1 and Day2'].index.to_list()[0]+2,
                          self.__plandf[self.__plandf['action']=='Break between Day2 and Day3'].index.to_list()[0]+2,
                          self.__plandf[self.__plandf['action']=='End of Day3'].index.to_list()[0]+2,]
        return None
    def __find_point(self,X,Y):
        eps=1/(10**self.__decimal)/2
        ptno=self.__pointdf[(self.__pointdf['X']>=X-eps) & (self.__pointdf['X']<X+eps) & (self.__pointdf['Y']>=Y-eps) & (self.__pointdf['Y']<Y+eps)]['No'].values[0]
        return ptno
    def write_excel(self):
        with pd.ExcelWriter(self.__outputpath) as writer:
            self.__plandf.to_excel(writer,index=False)
            worksheet=writer.sheets[list(writer.sheets.keys())[0]]
            for column_cells in worksheet.columns:
                lengthlist=[]
                for cell in column_cells:
                    length=0
                    for char in str(cell.value):
                        if char>=u'\u4e00' and char<=u'\u9fff':
                            length+=2
                        else:
                            length+=1
                    lengthlist.append(length)
                length=max(max(lengthlist),5)
                worksheet.column_dimensions[column_cells[0].column_letter].width=length+2
            for row in self.__voidindex:
                for cell in worksheet[row]:
                    cell.fill=PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
        timecost=self.__format_number(time.time()-self.__starttime)
        print('Process finished in '+timecost+' seconds')
        return None
    def schedule_frame(self):
        return self.__plandf.copy()
    def run(self):
        self.test_IO()
        self.read_info()
        self.read_task()
        self.read_package()
        self.read_point()
        self.gen_void_point()
        self.check_remote()
        self.divide_task()
        self.drop_O()
        self.update_matrix()
        self.read_schedule()
        if self.__write_output:
            self.write_excel()
        return None

if __name__=='__main__':
    opt=heuristic(CONST.NORMAL,'instance1')


def solve(case, mode="normal"):
    """Run the migrated heuristic algorithm on UnifiedCase input."""
    from .schedule import (
        SchedulePlan,
        build_legacy_bundle,
        legacy_schedule_to_rows,
    )

    mode_map = {
        CONST.NORMAL: CONST.NORMAL,
        CONST.REVISIONAL: CONST.REVISIONAL,
        CONST.BACK: CONST.BACK,
    }
    if mode not in mode_map:
        raise ValueError(f"Unsupported ha mode: {mode}")

    bundle = build_legacy_bundle(case)
    kwargs = {
        "situation": mode_map[mode],
        "decimal": case.config.algorithm.decimal,
        "writeOutput": False,
        "dataFrames": {
            "info": bundle.info,
            "task": bundle.task,
            "package": bundle.package[["name", "time", "power", "tag"]],
            "point": bundle.point[["name", "X", "Y", "备注"]],
            "distance": bundle.distance,
            "time": bundle.time,
            "power": bundle.power,
        },
    }
    if mode in {CONST.REVISIONAL, CONST.BACK}:
        baseline = case.config.raw.get("baseline_schedule")
        if not baseline:
            raise NotImplementedError(
                f"ha mode={mode} requires config.baseline_schedule"
            )
        baseline_path = (case.case_dir / baseline).resolve()
        if not baseline_path.exists():
            raise FileNotFoundError(f"Missing baseline schedule: {baseline_path}")
        kwargs["scheduleFrame"] = pd.read_csv(baseline_path)

    optimizer = heuristic(**kwargs)
    optimizer.run()
    schedule_df = optimizer.schedule_frame()
    rows = legacy_schedule_to_rows(case, schedule_df)
    objective_value = float(sum(row.get("revenue", 0) or 0 for row in rows))
    return SchedulePlan(steps=[], rows=rows, objective_value=objective_value)
